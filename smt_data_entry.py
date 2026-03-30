import sqlite3
import os
import re
import uuid
import json
import calendar
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Suggestion lists (used as default combo options)
# ---------------------------------------------------------------------------

PRODUCT_SUGGESTIONS = [
    "CDA67", "CDA67 TOP", "CDA67 BOT",
    "CDD41", "CDD41 TOP", "CDD41 BOT",
    "CDD21", "CDD21 TOP", "CDD21 BOT",
    "CDC10", "CDC10 TOP", "CDC10 BOT",
    "CDC16", "CDC22",
    "IFP 11", "IFP 32", "IFP 43", "IFP 55", "IFP 65",
    "OTHER",
]

FAULT_SUGGESTIONS = [
    "MISSING", "SHIFT", "TOMBSTONE", "FLIP", "DRY",
    "LIFTED", "LESS SOLDER", "EXTRA SOLDER", "BILL BOARD", "SHORT", "WRONG PART", "OTHER",
]

STAGE_SUGGESTIONS = ["AOI", "ICT", "FCT", "VISUAL", "OTHER"]
ACTION_SUGGESTIONS = ["REPAIRED", "REJECTED", "REWORK", "SCRAPED"]

# ---------------------------------------------------------------------------
# Default column definitions
# Each column: key (internal dict key), display (header label),
#              type ("entry" or "combo"), options (list for combo),
#              required (bool), default (str default value)
# ---------------------------------------------------------------------------

DEFAULT_COLUMNS = [
    {"key": "date",    "display": "DATE",                       "type": "entry", "options": [],                   "required": False, "default": ""},
    {"key": "product", "display": "PRODUCT",                    "type": "entry", "options": PRODUCT_SUGGESTIONS,  "required": False, "default": ""},
    {"key": "pcba_no", "display": "PCBA NO.",                   "type": "entry", "options": [],                   "required": True,  "default": ""},
    {"key": "problem", "display": "PROBLEM REPORTED",           "type": "entry", "options": [],                   "required": False, "default": ""},
    {"key": "fault",   "display": "FAULT",                      "type": "combo", "options": FAULT_SUGGESTIONS,    "required": False, "default": ""},
    {"key": "stage",   "display": "STAGE OF FAULT OBSERVATION", "type": "combo", "options": STAGE_SUGGESTIONS,    "required": False, "default": "AOI"},
    {"key": "action",  "display": "ACTION TAKEN",               "type": "combo", "options": ACTION_SUGGESTIONS,   "required": False, "default": "REPAIRED"},
]

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

import sys as _sys
_APP_DIR = (os.path.dirname(_sys.executable)
            if getattr(_sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__)))
COLUMNS_PATH = os.path.join(_APP_DIR, "columns.json")
DB_PATH      = os.path.join(_APP_DIR, "smt_rework.db")
QS_DB_PATH   = os.path.join(_APP_DIR, "smt_quality.db")

MONTH_NAMES  = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

# ---------------------------------------------------------------------------
# Column config persistence
# ---------------------------------------------------------------------------

def load_columns():
    """Load column definitions from columns.json, falling back to defaults."""
    if os.path.exists(COLUMNS_PATH):
        try:
            with open(COLUMNS_PATH, encoding="utf-8") as f:
                cols = json.load(f)
            for c in cols:
                c.setdefault("options",  [])
                c.setdefault("required", False)
                c.setdefault("default",  "")
                c.setdefault("type",     "entry")
            return cols
        except Exception:
            pass
    return [dict(c) for c in DEFAULT_COLUMNS]


def save_columns_cfg(columns):
    """Persist column definitions to columns.json."""
    with open(COLUMNS_PATH, "w", encoding="utf-8") as f:
        json.dump(columns, f, indent=2, ensure_ascii=False)

# ---------------------------------------------------------------------------
# Database setup
# ---------------------------------------------------------------------------

def get_connection():
    return sqlite3.connect(DB_PATH)


def init_db():
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                session_id  TEXT PRIMARY KEY,
                label       TEXT NOT NULL,
                created_at  TEXT NOT NULL
            )
        """)
        # Flexible schema: each entry's field data is stored as a JSON string
        conn.execute("""
            CREATE TABLE IF NOT EXISTS entries (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id  TEXT    NOT NULL,
                data        TEXT    NOT NULL,
                created_at  TEXT    NOT NULL,
                FOREIGN KEY (session_id) REFERENCES sessions(session_id)
            )
        """)
        conn.commit()
        _migrate_old_schema(conn)


def _migrate_old_schema(conn):
    """Migrate old fixed-column entries table to new JSON data format."""
    info      = conn.execute("PRAGMA table_info(entries)").fetchall()
    col_names = [row[1] for row in info]
    if "data" in col_names:
        return  # Already new format
    if "product" not in col_names:
        return  # No old schema to migrate

    # Rename old, recreate, migrate, drop old
    conn.execute("ALTER TABLE entries RENAME TO entries_old")
    conn.execute("""
        CREATE TABLE entries (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT    NOT NULL,
            data        TEXT    NOT NULL,
            created_at  TEXT    NOT NULL,
            FOREIGN KEY (session_id) REFERENCES sessions(session_id)
        )
    """)
    old_rows = conn.execute(
        "SELECT session_id, date, product, pcba_no, problem, fault, stage, action, created_at "
        "FROM entries_old"
    ).fetchall()
    for (session_id, d, product, pcba_no, problem, fault, stage, action, created_at) in old_rows:
        payload = json.dumps({
            "date":    d       or "",
            "product": product or "",
            "pcba_no": pcba_no or "",
            "problem": problem or "",
            "fault":   fault   or "",
            "stage":   stage   or "",
            "action":  action  or "",
        }, ensure_ascii=False)
        conn.execute(
            "INSERT INTO entries (session_id, data, created_at) VALUES (?, ?, ?)",
            (session_id, payload, created_at),
        )
    conn.execute("DROP TABLE entries_old")
    conn.commit()


_META_KEYS = {"id", "created_at", "session_id", "session_label", "session"}


def _row_to_entry(row_id, data_str, created_at, session_id=None, session_label=None):
    try:
        entry = json.loads(data_str)
    except Exception:
        entry = {}
    entry["id"]         = row_id
    entry["created_at"] = created_at
    if session_id    is not None: entry["session_id"]    = session_id
    if session_label is not None: entry["session_label"] = session_label
    return entry


def save_entries(entries, session_id, session_label):
    now = datetime.now().isoformat(timespec="seconds")
    with get_connection() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO sessions (session_id, label, created_at) VALUES (?, ?, ?)",
            (session_id, session_label, now),
        )
        for entry in entries:
            data = {k: v for k, v in entry.items() if k not in _META_KEYS}
            conn.execute(
                "INSERT INTO entries (session_id, data, created_at) VALUES (?, ?, ?)",
                (session_id, json.dumps(data, ensure_ascii=False), now),
            )
        conn.commit()


def search_entries(date_from=None, date_to=None, text_search=None):
    """Return all entries matching the supplied filters (Python-side filtering)."""
    with get_connection() as conn:
        rows = conn.execute("""
            SELECT e.id, e.data, e.created_at, e.session_id, s.label AS session_label
            FROM entries e
            JOIN sessions s ON e.session_id = s.session_id
            ORDER BY e.id DESC
        """).fetchall()

    results = []
    for (row_id, data_str, created_at, session_id, session_label) in rows:
        entry = _row_to_entry(row_id, data_str, created_at, session_id, session_label)
        date_val = entry.get("date", "")
        if date_from and date_val < date_from: continue
        if date_to   and date_val > date_to:   continue
        if text_search:
            # Match whole "token" — not embedded inside another number.
            # e.g. "901" matches "R901" or "901A" but NOT "90159".
            pattern = re.compile(r'(?<!\d)' + re.escape(text_search) + r'(?!\d)',
                                 re.IGNORECASE)
            if not any(pattern.search(str(v))
                       for k, v in entry.items()
                       if k not in _META_KEYS and v):
                continue
        results.append(entry)
    return results


def get_all_model_values():
    """Return sorted list of all unique model/product values stored in the DB."""
    with get_connection() as conn:
        rows = conn.execute("SELECT data FROM entries").fetchall()
    models = set()
    for (data_json,) in rows:
        try:
            d = json.loads(data_json)
            for k, v in d.items():
                if any(x in k.lower() for x in ("model", "product")) and v:
                    models.add(str(v))
        except Exception:
            pass
    return sorted(models)

def get_all_sessions():
    with get_connection() as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT session_id, label, created_at FROM sessions ORDER BY created_at DESC"
        ).fetchall()
    return [dict(row) for row in rows]


def get_session_entries(session_id):
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT id, data, created_at FROM entries WHERE session_id = ? ORDER BY id ASC",
            (session_id,),
        ).fetchall()
    return [_row_to_entry(r[0], r[1], r[2]) for r in rows]


def update_db_entry(entry_id, entry):
    data = {k: v for k, v in entry.items() if k not in _META_KEYS}
    with get_connection() as conn:
        conn.execute(
            "UPDATE entries SET data = ? WHERE id = ?",
            (json.dumps(data, ensure_ascii=False), entry_id),
        )
        conn.commit()


def delete_db_entry(entry_id):
    with get_connection() as conn:
        conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
        conn.commit()


def delete_db_session(session_id):
    with get_connection() as conn:
        conn.execute("DELETE FROM entries  WHERE session_id = ?", (session_id,))
        conn.execute("DELETE FROM sessions WHERE session_id = ?", (session_id,))
        conn.commit()

def rename_db_session(session_id, new_label):
    with get_connection() as conn:
        conn.execute("UPDATE sessions SET label = ? WHERE session_id = ?",
                     (new_label, session_id))
        conn.commit()

# ---------------------------------------------------------------------------
# Excel export  (uses dynamic column list)
# ---------------------------------------------------------------------------

def export_to_excel(entries, filepath, columns):
    from collections import Counter
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    DARK_BLUE  = "1F4E79"
    LIGHT_BLUE = "DEEAF1"
    WHITE      = "FFFFFF"

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Rework Register"

    HEADERS = [c["display"] for c in columns]
    KEYS    = [c["key"]     for c in columns]

    hdr_font  = Font(bold=True, color=WHITE)
    hdr_fill  = PatternFill("solid", fgColor=DARK_BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    row_align = Alignment(vertical="center")

    for col, text in enumerate(HEADERS, 1):
        cell           = ws1.cell(row=1, column=col, value=text)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws1.row_dimensions[1].height = 30

    for row_idx, entry in enumerate(entries, 2):
        for col, key in enumerate(KEYS, 1):
            cell           = ws1.cell(row=row_idx, column=col, value=entry.get(key, ""))
            cell.alignment = row_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, _ in enumerate(KEYS, 1):
        letter  = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws1.cell(row=r, column=col_idx).value or ""))
             for r in range(1, len(entries) + 2)),
            default=0,
        )
        ws1.column_dimensions[letter].width = max(10, max_len + 3)

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Chart Data ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chart Data")

    def _write_table(ws, start_col, hdr_a, hdr_b, rows):
        for offset, hdr in enumerate((hdr_a, hdr_b)):
            c           = ws.cell(row=1, column=start_col + offset, value=hdr)
            c.font      = Font(bold=True, color=WHITE)
            c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
            c.alignment = Alignment(horizontal="center")
        for r, (a, b) in enumerate(rows, 2):
            ws.cell(row=r, column=start_col,     value=a)
            ws.cell(row=r, column=start_col + 1, value=b)

    from openpyxl.styles import Border, Side

    fault_key   = next((c["key"] for c in columns if "fault"   in c["key"].lower()), None)
    date_key    = next((c["key"] for c in columns if "date"    in c["key"].lower()), None)
    problem_key = next((c["key"] for c in columns if "problem" in c["key"].lower()), None)

    col_offset = 1
    summary_rows = 1  # track deepest row written in the summary section
    if fault_key:
        fc = Counter(e.get(fault_key, "") or "" for e in entries)
        rows_fc = sorted(((k, v) for k, v in fc.items() if k), key=lambda x: x[1], reverse=True)
        _write_table(ws2, col_offset, "FAULT TYPE", "COUNT", rows_fc)
        summary_rows = max(summary_rows, 1 + len(rows_fc))
        col_offset += 3
    if date_key:
        dc = Counter(e.get(date_key, "") or "" for e in entries)
        rows_dc = sorted(((k, v) for k, v in dc.items() if k), key=lambda x: x[0])
        _write_table(ws2, col_offset, "DATE", "COUNT", rows_dc)
        summary_rows = max(summary_rows, 1 + len(rows_dc))
        col_offset += 3
    if problem_key:
        pc = Counter(e.get(problem_key, "") or "" for e in entries)
        rows_pc = sorted(((k, v) for k, v in pc.items() if k), key=lambda x: x[1], reverse=True)
        _write_table(ws2, col_offset, "COMPONENT", "COUNT", rows_pc)
        summary_rows = max(summary_rows, 1 + len(rows_pc))

    # ── Breakdown tables ──────────────────────────────────────────────────────
    ORANGE  = "C25700"   # dark orange for TOTAL row (matches screenshot)
    thin_s  = Side(style="thin", color="BBBBBB")
    thin_b  = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    def _write_breakdown(ws, start_row, start_col, title, items, pct_label):
        """3-column block: name | COUNT | %.  Returns next free row after block."""
        hdr_fill  = PatternFill("solid", fgColor=DARK_BLUE)
        tot_fill  = PatternFill("solid", fgColor=ORANGE)
        ctr       = Alignment(horizontal="center", vertical="center")
        hdr_font  = Font(bold=True, color=WHITE, size=9)

        total_cnt = sum(cnt for _, cnt in items)

        # Title cell (spans 3 cols)
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row, end_column=start_col + 2)
        tc = ws.cell(start_row, start_col, title)
        tc.font      = hdr_font
        tc.fill      = hdr_fill
        tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sub-header row
        sh = start_row + 1
        for off, txt in enumerate(["", "COUNT", pct_label]):
            c = ws.cell(sh, start_col + off, txt)
            c.font      = Font(bold=True, color=WHITE, size=8)
            c.fill      = hdr_fill
            c.alignment = ctr

        # Data rows (all items, sorted desc already)
        r = sh + 1
        alt_fill2 = PatternFill("solid", fgColor="FFF0E6")
        for idx, (name, cnt) in enumerate(items):
            pct  = f"{cnt / total_cnt * 100:.1f}%" if total_cnt else "0.0%"
            fill = alt_fill2 if idx % 2 == 1 else PatternFill()
            for off, val in enumerate([name, cnt, pct]):
                c = ws.cell(r, start_col + off, val)
                c.alignment = ctr if off > 0 else Alignment(vertical="center")
                c.border    = thin_b
                if fill.fill_type:
                    c.fill = fill
            r += 1

        # TOTAL row
        tot_pct = "100.0%" if total_cnt else "0.0%"
        for off, val in enumerate(["TOTAL", total_cnt, tot_pct]):
            c = ws.cell(r, start_col + off, val)
            c.font      = Font(bold=True, color=WHITE, size=9)
            c.fill      = tot_fill
            c.alignment = ctr
            c.border    = thin_b

        return r + 1   # next free row

    if fault_key and problem_key:
        # ── Section A: top-5 defects → component breakdown ───────────────────
        all_faults = Counter(e.get(fault_key, "") or "" for e in entries)
        top_faults = [f for f, _ in all_faults.most_common() if f][:5]

        section_start = summary_rows + 3   # gap below summary
        # Section label
        ws2.merge_cells(start_row=section_start - 1, start_column=1,
                        end_row=section_start - 1, end_column=4 * len(top_faults))
        lbl = ws2.cell(section_start - 1, 1,
                       "DEFECT-WISE COMPONENT BREAKDOWN (TOP 5 DEFECTS)")
        lbl.font      = Font(bold=True, color=WHITE, size=10)
        lbl.fill      = PatternFill("solid", fgColor="1F4E79")
        lbl.alignment = Alignment(horizontal="center")

        breakdown_depths = []
        for t_idx, fault in enumerate(top_faults):
            components = Counter(
                e.get(problem_key, "") or ""
                for e in entries
                if (e.get(fault_key, "") or "") == fault
            )
            items = [(k, v) for k, v in components.most_common() if k]
            sc = 1 + t_idx * 4
            depth = _write_breakdown(ws2, section_start, sc,
                                     f"{fault.upper()} – COMPONENT BREAKDOWN",
                                     items, "OF FAULT %")
            breakdown_depths.append(depth)

        # ── Section B: top-5 components → defect breakdown ───────────────────
        all_comps = Counter(e.get(problem_key, "") or "" for e in entries)
        top_comps = [c for c, _ in all_comps.most_common() if c][:5]

        section2_start = max(breakdown_depths) + 2 if breakdown_depths else section_start + 10
        ws2.merge_cells(start_row=section2_start - 1, start_column=1,
                        end_row=section2_start - 1, end_column=4 * len(top_comps))
        lbl2 = ws2.cell(section2_start - 1, 1,
                        "COMPONENT-WISE DEFECT BREAKDOWN (TOP 5 COMPONENTS)")
        lbl2.font      = Font(bold=True, color=WHITE, size=10)
        lbl2.fill      = PatternFill("solid", fgColor="1F4E79")
        lbl2.alignment = Alignment(horizontal="center")

        for t_idx, comp in enumerate(top_comps):
            defects = Counter(
                e.get(fault_key, "") or ""
                for e in entries
                if (e.get(problem_key, "") or "") == comp
            )
            items = [(k, v) for k, v in defects.most_common() if k]
            sc = 1 + t_idx * 4
            _write_breakdown(ws2, section2_start, sc,
                             f"{comp} – DEFECT BREAKDOWN",
                             items, "OF COMPONENT %")

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {}
    for row in ws2.iter_rows():
        for c in row:
            if c.value and not getattr(c, "data_type", None) == "n":
                cw = col_widths.get(c.column, 8)
                col_widths[c.column] = max(cw, len(str(c.value)) + 3)
    for col_num, width in col_widths.items():
        ws2.column_dimensions[get_column_letter(col_num)].width = min(width, 30)

    wb.save(filepath)

# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------

class SMTReworkApp:
    DARK_BLUE = "#1F4E79"
    MED_BLUE  = "#2E75B6"
    GREEN     = "#217346"
    BG        = "#F0F4F8"
    CARD_BG   = "#FFFFFF"
    ALT_ROW   = "#DEEAF1"

    def __init__(self, root):
        self.root = root
        self.root.title("SMT Production Data Entry System")
        self.root.geometry("1100x720")
        self.root.minsize(900, 600)
        self.root.configure(bg=self.BG)

        self.columns                = load_columns()
        self.session_entries        = []
        self.session_id             = str(uuid.uuid4())
        self.session_saved_count    = 0
        self._action_pending        = False
        self._hist_action_pending   = False
        self._hist_current_results  = []
        self._hist_raw_results      = []
        self.hist_combo_filters     = []   # list of (key, StringVar) for history filter row2
        self._hist_effective_columns = self.columns  # columns currently shown in hist treeview

        # form widget references (rebuilt dynamically)
        self.col_vars    = {}   # key → StringVar
        self.col_widgets = {}   # key → widget

        init_db()
        qs_init_db()
        self._setup_styles()
        self._build_ui()
        self._focus_key_widget()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── styles ────────────────────────────────────────────────────────────────

    def _setup_styles(self):
        s = ttk.Style(self.root)
        s.theme_use("clam")

        s.configure("TFrame",        background=self.BG)
        s.configure("TLabel",        background=self.BG,    font=("Segoe UI", 10))
        s.configure("TEntry",        font=("Segoe UI", 10), padding=(4, 3))
        s.configure("TCombobox",     font=("Segoe UI", 10), padding=(4, 3))
        s.configure("TNotebook",     background=self.BG)
        s.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"),
                    padding=(16, 6), background="#C8D8EA")
        s.map("TNotebook.Tab",
              background=[("selected", self.DARK_BLUE)],
              foreground=[("selected", "white")])

        s.configure("Card.TFrame", background=self.CARD_BG)

        s.configure("Dark.TButton",
                    background=self.DARK_BLUE, foreground="white",
                    font=("Segoe UI", 10, "bold"), relief="flat", borderwidth=0)
        s.map("Dark.TButton",
              background=[("active", self.MED_BLUE), ("pressed", "#163D61")])

        s.configure("Green.TButton",
                    background=self.GREEN, foreground="white",
                    font=("Segoe UI", 10, "bold"), relief="flat", borderwidth=0)
        s.map("Green.TButton",
              background=[("active", "#1A5C38"), ("pressed", "#134429")])

        s.configure("Muted.TButton",
                    background="#C8D4E0", foreground="#333333",
                    font=("Segoe UI", 10), relief="flat", borderwidth=0)
        s.map("Muted.TButton", background=[("active", "#B0BFD0")])

        s.configure("Danger.TButton",
                    background="#C0392B", foreground="white",
                    font=("Segoe UI", 10), relief="flat", borderwidth=0)
        s.map("Danger.TButton",
              background=[("active", "#E74C3C"), ("pressed", "#922B21")])

        s.configure("Treeview",
                    background=self.CARD_BG, foreground="#1A1A1A",
                    rowheight=26, fieldbackground=self.CARD_BG,
                    font=("Segoe UI", 9), borderwidth=0)
        s.configure("Treeview.Heading",
                    background=self.DARK_BLUE, foreground="white",
                    font=("Segoe UI", 9, "bold"), relief="flat")
        s.map("Treeview.Heading",
              background=[("active", self.MED_BLUE)])
        s.map("Treeview",
              background=[("selected", "#BDD7EE")],
              foreground=[("selected", "#000000")])

    # ── autocomplete helper ───────────────────────────────────────────────────

    @staticmethod
    def _bind_autocomplete(cb, all_values, sv):
        def _on_key(event):
            nav = ('BackSpace', 'Delete', 'Left', 'Right', 'Up', 'Down',
                   'Home', 'End', 'Escape', 'Tab', 'Return', 'F4',
                   'Prior', 'Next', 'Shift_L', 'Shift_R', 'Control_L',
                   'Control_R', 'Alt_L', 'Alt_R')
            typed = sv.get()
            if event.keysym in nav:
                cb["values"] = (
                    [v for v in all_values if v.upper().startswith(typed.upper())]
                    or all_values
                )
                return
            if not typed:
                cb["values"] = all_values
                return
            filtered = [v for v in all_values if v.upper().startswith(typed.upper())]
            cb["values"] = filtered or all_values
            if filtered:
                sv.set(filtered[0])
                cb.icursor(len(typed))
                try:
                    cb.selection_range(len(typed), "end")
                except tk.TclError:
                    pass
        cb.bind("<KeyRelease>", _on_key)

    # ── top-level layout ──────────────────────────────────────────────────────

    def _build_ui(self):
        tb = tk.Frame(self.root, bg=self.DARK_BLUE, height=52)
        tb.pack(fill="x")
        tb.pack_propagate(False)
        tk.Label(tb, text="SMT Production Data Entry System",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(tb, text="SMT Department",
                 bg=self.DARK_BLUE, fg="#9DB8D8",
                 font=("Segoe UI", 10)).pack(side="right", padx=20)

        # Outer notebook: Rework Register | Quality Sheet
        self.outer_nb = ttk.Notebook(self.root)
        self.outer_nb.pack(fill="both", expand=True)

        rework_tab  = ttk.Frame(self.outer_nb)
        self.outer_nb.add(rework_tab,  text="  Rework Register  ")

        quality_tab = ttk.Frame(self.outer_nb)
        self.outer_nb.add(quality_tab, text="  Quality Sheet  ")

        # Rework Register: inner notebook (Data Entry + History)
        self.notebook = ttk.Notebook(rework_tab)
        self.notebook.pack(fill="both", expand=True)

        tab1 = ttk.Frame(self.notebook)
        self.notebook.add(tab1, text="  Data Entry  ")

        tab2 = ttk.Frame(self.notebook)
        self.notebook.add(tab2, text="  History & Search  ")

        self._build_data_entry_tab(tab1)
        self._build_history_tab(tab2)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

        # Quality Sheet tab
        self._qs_frame = QualitySheetFrame(quality_tab, self)
        self._qs_frame.pack(fill="both", expand=True)


    # ── Data Entry tab ────────────────────────────────────────────────────────

    def _build_data_entry_tab(self, parent):
        self._build_bottom_bar(parent)

        content = ttk.Frame(parent)
        content.pack(fill="both", expand=True, padx=10, pady=(10, 6))

        left = tk.Frame(content, bg=self.CARD_BG, width=292)
        left.pack(side="left", fill="y", padx=(0, 8))
        left.pack_propagate(False)
        self._build_form(left)

        right = tk.Frame(content, bg=self.CARD_BG)
        right.pack(side="left", fill="both", expand=True)
        self._build_table(right)

    def _build_bottom_bar(self, parent):
        bar = tk.Frame(parent, bg=self.BG, height=52)
        bar.pack(side="bottom", fill="x", padx=10, pady=(0, 8))
        bar.pack_propagate(False)

        ttk.Button(bar, text="Export to Excel",
                   style="Green.TButton",
                   command=self.export_session
                   ).pack(side="left", pady=10, ipady=5, ipadx=14)

        ttk.Button(bar, text="Import Excel",
                   style="Muted.TButton",
                   command=self.import_excel
                   ).pack(side="left", padx=6, pady=10, ipady=5, ipadx=10)

        ttk.Button(bar, text="Save Session",
                   style="Dark.TButton",
                   command=self.save_session
                   ).pack(side="left", padx=(2, 6), pady=10, ipady=5, ipadx=14)

        ttk.Button(bar, text="\u21a9 Undo Last",
                   style="Muted.TButton",
                   command=self._undo_last_entry
                   ).pack(side="left", pady=10, ipady=5, ipadx=10)

        ttk.Button(bar, text="Find & Replace",
                   style="Muted.TButton",
                   command=self._show_find_replace
                   ).pack(side="left", padx=6, pady=10, ipady=5, ipadx=10)

        ttk.Button(bar, text="\u2630 Manage Columns",
                   style="Muted.TButton",
                   command=self._show_manage_columns
                   ).pack(side="left", pady=10, ipady=5, ipadx=10)

        ttk.Button(bar, text="Clear Session",
                   style="Muted.TButton",
                   command=self.clear_session
                   ).pack(side="left", padx=6, pady=10, ipady=5, ipadx=10)

        self.lbl_session_info = tk.Label(
            bar,
            text=f"Session: {self.session_id[:8].upper()}",
            bg=self.BG, fg="#9090A0", font=("Segoe UI", 9))
        self.lbl_session_info.pack(side="right", padx=6)


    # ── form (left panel) ─────────────────────────────────────────────────────

    def _build_form(self, parent):
        hdr = tk.Frame(parent, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="New Entry",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=10)

        # Scrollable fields area
        scroll_outer = tk.Frame(parent, bg=self.CARD_BG)
        scroll_outer.pack(fill="both", expand=True)

        self._form_canvas = tk.Canvas(scroll_outer, bg=self.CARD_BG, highlightthickness=0)
        form_vsb = ttk.Scrollbar(scroll_outer, orient="vertical", command=self._form_canvas.yview)
        self._form_canvas.configure(yscrollcommand=form_vsb.set)
        form_vsb.pack(side="right", fill="y")
        self._form_canvas.pack(side="left", fill="both", expand=True)

        self.form_inner = tk.Frame(self._form_canvas, bg=self.CARD_BG, padx=14, pady=8)
        self._form_win  = self._form_canvas.create_window((0, 0), window=self.form_inner, anchor="nw")

        self.form_inner.bind("<Configure>",
            lambda e: self._form_canvas.configure(scrollregion=self._form_canvas.bbox("all")))
        self._form_canvas.bind("<Configure>",
            lambda e: self._form_canvas.itemconfig(self._form_win, width=e.width))

        def _on_mw(event):
            self._form_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._form_canvas.bind("<Enter>",  lambda e: self._form_canvas.bind_all("<MouseWheel>", _on_mw))
        self._form_canvas.bind("<Leave>",  lambda e: self._form_canvas.unbind_all("<MouseWheel>"))

        self._populate_form_fields()

        # Fixed bottom section (buttons + counter)
        bottom = tk.Frame(parent, bg=self.CARD_BG, padx=14, pady=8)
        bottom.pack(fill="x", side="bottom")

        tk.Frame(bottom, bg="#DDE6EF", height=1).pack(fill="x", pady=(0, 8))
        ttk.Button(bottom, text="Add Entry  \u21b5",
                   style="Dark.TButton",
                   command=self._on_add_entry).pack(fill="x", ipady=8)
        tk.Frame(bottom, bg=self.CARD_BG, height=5).pack()
        ttk.Button(bottom, text="Clear Form",
                   style="Muted.TButton",
                   command=self._clear_form).pack(fill="x", ipady=5)
        tk.Frame(bottom, bg="#DDE6EF", height=1).pack(fill="x", pady=(8, 4))
        self.lbl_counter = tk.Label(
            bottom, text="Entries this session:  0",
            bg=self.CARD_BG, fg=self.DARK_BLUE,
            font=("Segoe UI", 10, "bold"))
        self.lbl_counter.pack(anchor="w")

    def _populate_form_fields(self):
        """Build / rebuild form fields from self.columns."""
        for w in self.form_inner.winfo_children():
            w.destroy()
        self.col_vars    = {}
        self.col_widgets = {}

        today = date.today().strftime("%Y-%m-%d")

        for col_def in self.columns:
            key     = col_def["key"]
            display = col_def["display"]
            typ     = col_def.get("type", "entry")
            opts    = col_def.get("options", [])
            default = col_def.get("default", "")

            # Use today's date as default for date-like columns
            if "date" in key.lower() and not default:
                default = today

            sv = tk.StringVar(value=default)
            self.col_vars[key] = sv

            lbl_text = display + ("  \u2605" if col_def.get("required") else "")
            tk.Label(self.form_inner, text=lbl_text, bg=self.CARD_BG,
                     fg="#555555", font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 1))

            if typ == "combo" and opts:
                w = ttk.Combobox(self.form_inner, textvariable=sv, values=opts,
                                 font=("Segoe UI", 10))
                w.pack(fill="x", ipady=2)
                self._bind_autocomplete(w, opts, sv)
            else:
                big = col_def.get("required", False)
                w = ttk.Entry(self.form_inner, textvariable=sv,
                              font=("Segoe UI", 11 if big else 10))
                w.pack(fill="x", ipady=3)
            self.col_widgets[key] = w

        # Enter-key focus chain
        keys = [c["key"] for c in self.columns]
        for i, key in enumerate(keys[:-1]):
            nxt = keys[i + 1]
            self.col_widgets[key].bind(
                "<Return>", lambda e, n=nxt: self.col_widgets[n].focus_set())
        if keys:
            self.col_widgets[keys[-1]].bind("<Return>", self._on_add_entry)

    def _focus_key_widget(self):
        """Focus the first required field, or the first field."""
        for col_def in self.columns:
            if col_def.get("required"):
                w = self.col_widgets.get(col_def["key"])
                if w:
                    w.focus_set()
                    return
        if self.col_widgets:
            first = self.columns[0]["key"] if self.columns else None
            if first and first in self.col_widgets:
                self.col_widgets[first].focus_set()

    # ── table (right panel) ───────────────────────────────────────────────────

    def _build_table(self, parent):
        hdr = tk.Frame(parent, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Current Session Entries",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=10)
        tk.Label(hdr, text="double-click a row to edit / delete",
                 bg=self.DARK_BLUE, fg="#9DB8D8",
                 font=("Segoe UI", 9)).pack(side="right", padx=14)

        self.sv_sess_model_filter = tk.StringVar(value="All")
        self.sv_sess_model_filter.trace_add("write", lambda *_: self._refresh_tree())

        self.tv_frame = tk.Frame(parent, bg=self.CARD_BG)
        self.tv_frame.pack(fill="both", expand=True)
        self._create_treeview()

        # ── row-reorder / insert toolbar ──────────────────────────────────
        row_bar = tk.Frame(parent, bg=self.BG, pady=3)
        row_bar.pack(fill="x", padx=4)
        ttk.Button(row_bar, text="\u2191 Move Row Up",
                   style="Muted.TButton",
                   command=self._row_move_up
                   ).pack(side="left", padx=(0, 4), ipady=3, ipadx=8)
        ttk.Button(row_bar, text="\u2193 Move Row Down",
                   style="Muted.TButton",
                   command=self._row_move_down
                   ).pack(side="left", padx=(0, 4), ipady=3, ipadx=8)
        ttk.Button(row_bar, text="+ Insert Below Selected",
                   style="Dark.TButton",
                   command=self._insert_below_selected
                   ).pack(side="left", ipady=3, ipadx=8)
        self.var_sort_by_date = tk.BooleanVar(value=True)
        _DATE_FORMATS = ["YYYY-MM-DD", "DD-MM-YYYY", "DD/MM/YYYY",
                         "MM-DD-YYYY", "MM/DD/YYYY", "DD.MM.YYYY",
                         "DD.MM.YY",  "DD-MM-YY",   "DD/MM/YY"]
        self.var_date_fmt = tk.StringVar(value="DD.MM.YY")
        ttk.Checkbutton(row_bar, text="Sort by Date",
                        variable=self.var_sort_by_date,
                        command=self._refresh_tree
                        ).pack(side="right", padx=(4, 0))
        self.cmb_date_fmt = ttk.Combobox(row_bar, textvariable=self.var_date_fmt,
                                         values=_DATE_FORMATS, state="readonly", width=12)
        self.cmb_date_fmt.pack(side="right", padx=(0, 4))
        self.var_date_fmt.trace_add("write", lambda *_: self._refresh_tree())

    def _create_treeview(self):
        """Create / recreate the session treeview with current column layout."""
        for w in self.tv_frame.winfo_children():
            w.destroy()

        data_cols = tuple(c["key"] for c in self.columns)
        all_cols  = data_cols + ("edit_col", "del_col")

        self.tree = ttk.Treeview(self.tv_frame, columns=all_cols,
                                 show="headings", selectmode="browse")

        model_key = self._model_col_key()
        for col_def in self.columns:
            key  = col_def["key"]
            disp = col_def["display"]
            w    = max(70, min(200, len(disp) * 9 + 16))
            if key == model_key:
                self.tree.heading(key, text=f"{disp} ▼",
                                  command=self._show_sess_model_popup)
            else:
                self.tree.heading(key, text=disp)
            self.tree.column(key, width=w, anchor="center", minwidth=50)

        self.tree.heading("edit_col", text="\u270f")
        self.tree.column("edit_col", width=34, anchor="center", minwidth=34)
        self.tree.heading("del_col", text="\U0001f5d1")
        self.tree.column("del_col", width=34, anchor="center", minwidth=34)

        self.tree.tag_configure("odd",  background=self.CARD_BG)
        self.tree.tag_configure("even", background=self.ALT_ROW)

        vsb = ttk.Scrollbar(self.tv_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tv_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right",  fill="y")
        self.tree.pack(fill="both", expand=True)

        self.tree.bind("<Double-1>", self._on_row_double_click)
        self.tree.bind("<Button-1>",  self._on_tree_click)

        self._refresh_tree()

    # ── entry logic ───────────────────────────────────────────────────────────

    def _on_add_entry(self, event=None):
        # Validate required columns
        for col_def in self.columns:
            if col_def.get("required"):
                key = col_def["key"]
                val = self.col_vars[key].get().strip()
                if not val:
                    messagebox.showwarning("Missing Field",
                                          f"{col_def['display']} cannot be empty.")
                    self.col_widgets[key].focus_set()
                    return

        # Build entry dict
        entry = {}
        for col_def in self.columns:
            key = col_def["key"]
            val = self.col_vars[key].get().strip()
            if key == "pcba_no":
                val = val.upper()
            entry[key] = val

        # Duplicate check: warn if all field values match an existing entry
        keys = [c["key"] for c in self.columns]
        for existing in self.session_entries:
            if all(entry.get(k, "") == existing.get(k, "") for k in keys):
                if not messagebox.askyesno(
                        "Duplicate Entry",
                        "This entry is identical to an existing row in the session.\n"
                        "Add it anyway?"):
                    return
                break

        self.session_entries.append(entry)
        n = len(self.session_entries)
        self._refresh_tree()
        # Scroll to the newly added entry after sort
        date_key = next((c["key"] for c in self.columns if "date" in c["key"].lower()), None)
        new_val = entry.get(date_key, "") if date_key else ""
        new_idx = next((i for i, e in enumerate(self.session_entries)
                        if e is entry), n - 1)
        self.tree.see(str(new_idx))

        # Clear entry-type fields (except date-like); keep combo defaults
        for col_def in self.columns:
            key = col_def["key"]
            if col_def.get("type", "entry") != "combo" and "date" not in key.lower():
                self.col_vars[key].set("")

        self.lbl_counter.config(text=f"Entries this session:  {n}")
        self._focus_key_widget()

    def _clear_form(self):
        for col_def in self.columns:
            key = col_def["key"]
            if col_def.get("type", "entry") != "combo" and "date" not in key.lower():
                self.col_vars[key].set("")
        self._focus_key_widget()

    # ── row reorder / insert ──────────────────────────────────────────────────

    def _selected_row_idx(self):
        """Return the index of the currently selected treeview row, or None."""
        sel = self.tree.selection()
        if sel:
            return int(sel[0])
        return None

    def _row_move_up(self):
        idx = self._selected_row_idx()
        if idx is None or idx == 0:
            return
        self.session_entries[idx - 1], self.session_entries[idx] = \
            self.session_entries[idx], self.session_entries[idx - 1]
        self._refresh_tree()
        new_iid = str(idx - 1)
        self.tree.selection_set(new_iid)
        self.tree.see(new_iid)

    def _row_move_down(self):
        idx = self._selected_row_idx()
        if idx is None or idx >= len(self.session_entries) - 1:
            return
        self.session_entries[idx + 1], self.session_entries[idx] = \
            self.session_entries[idx], self.session_entries[idx + 1]
        self._refresh_tree()
        new_iid = str(idx + 1)
        self.tree.selection_set(new_iid)
        self.tree.see(new_iid)

    def _insert_below_selected(self):
        """Validate the form, build an entry, and insert it after the selected row.
        Falls back to appending if nothing is selected."""
        # Validate required columns
        for col_def in self.columns:
            if col_def.get("required"):
                key = col_def["key"]
                val = self.col_vars[key].get().strip()
                if not val:
                    messagebox.showwarning("Missing Field",
                                          f"{col_def['display']} cannot be empty.")
                    self.col_widgets[key].focus_set()
                    return
        # Build entry dict
        entry = {}
        for col_def in self.columns:
            key = col_def["key"]
            val = self.col_vars[key].get().strip()
            if key == "pcba_no":
                val = val.upper()
            entry[key] = val
        # Determine insertion position
        idx = self._selected_row_idx()
        if idx is None:
            self.session_entries.append(entry)
        else:
            self.session_entries.insert(idx + 1, entry)
        self._refresh_tree()
        # Select and scroll to the newly inserted row
        new_idx = (idx + 1) if idx is not None else len(self.session_entries) - 1
        new_iid = str(new_idx)
        self.tree.selection_set(new_iid)
        self.tree.see(new_iid)
        # Clear entry-type fields
        for col_def in self.columns:
            key = col_def["key"]
            if col_def.get("type", "entry") != "combo" and "date" not in key.lower():
                self.col_vars[key].set("")
        self.lbl_counter.config(
            text=f"Entries this session:  {len(self.session_entries)}")
        self._focus_key_widget()

    # ── row edit / delete ─────────────────────────────────────────────────────

    def _on_row_double_click(self, event):
        n = len(self.columns)
        col = self.tree.identify_column(event.x)
        if col in (f"#{n+1}", f"#{n+2}"):
            return
        item = self.tree.identify_row(event.y)
        if item:
            self._show_edit_dialog(int(item))

    def _show_edit_dialog(self, idx):
        entry = self.session_entries[idx]

        req_key  = next((c["key"] for c in self.columns if c.get("required")), "pcba_no")
        req_disp = entry.get(req_key, "Entry")

        dlg = tk.Toplevel(self.root)
        dlg.title("Edit / Delete Entry")
        dlg.geometry("440x460")
        dlg.resizable(True, True)
        dlg.configure(bg=self.BG)
        dlg.transient(self.root)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"Edit / Delete  \u2014  {req_disp}",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=14, pady=10)

        # Scrollable form
        frm_outer = tk.Frame(dlg, bg=self.CARD_BG)
        frm_outer.pack(fill="both", expand=True, padx=16, pady=8)

        canvas = tk.Canvas(frm_outer, bg=self.CARD_BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(frm_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        frm = tk.Frame(canvas, bg=self.CARD_BG)
        win = canvas.create_window((0, 0), window=frm, anchor="nw")
        frm.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        svars = {}
        for r, col_def in enumerate(self.columns):
            key  = col_def["key"]
            typ  = col_def.get("type", "entry")
            opts = col_def.get("options", [])
            sv   = tk.StringVar(value=entry.get(key, ""))
            svars[key] = sv

            tk.Label(frm, text=col_def["display"], bg=self.CARD_BG,
                     fg="#555555", font=("Segoe UI", 9)
                     ).grid(row=r, column=0, sticky="w", padx=(0, 10), pady=(4, 0))

            if typ == "combo" and opts:
                w = ttk.Combobox(frm, textvariable=sv, values=opts, width=26)
                self._bind_autocomplete(w, opts, sv)
            elif typ == "combo":
                w = ttk.Combobox(frm, textvariable=sv, width=26)
            else:
                w = ttk.Entry(frm, textvariable=sv, width=28)
            w.grid(row=r, column=1, sticky="ew", pady=(4, 0))
        frm.columnconfigure(1, weight=1)

        def _save():
            updated = {}
            for col_def in self.columns:
                key = col_def["key"]
                val = svars[key].get().strip()
                if key == "pcba_no":
                    val = val.upper()
                updated[key] = val
            self.session_entries[idx] = updated
            tree_vals = [updated.get(c["key"], "") for c in self.columns] + ["\u270f", "\U0001f5d1"]
            self.tree.item(str(idx), values=tree_vals)
            dlg.destroy()

        def _delete():
            disp_val = entry.get(req_key, "this entry")
            if messagebox.askyesno("Confirm Delete",
                                   f"Delete entry for '{disp_val}'?", parent=dlg):
                self.session_entries.pop(idx)
                self._refresh_tree()
                self.lbl_counter.config(
                    text=f"Entries this session:  {len(self.session_entries)}")
                dlg.destroy()

        btn_bar = tk.Frame(dlg, bg=self.BG, padx=14, pady=10)
        btn_bar.pack(fill="x")
        ttk.Button(btn_bar, text="Save Changes", style="Dark.TButton",
                   command=_save).pack(side="left", expand=True, fill="x",
                                       padx=(0, 6), ipady=6)
        tk.Button(btn_bar, text="Delete",
                  bg="#C0392B", fg="white",
                  activebackground="#922B21", activeforeground="white",
                  font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                  command=_delete).pack(side="left", expand=True, fill="x",
                                        padx=(0, 6), ipady=8)
        ttk.Button(btn_bar, text="Cancel", style="Muted.TButton",
                   command=dlg.destroy).pack(side="left", expand=True,
                                              fill="x", ipady=6)

    def _sort_by_date(self):
        if not getattr(self, "var_sort_by_date", None) or not self.var_sort_by_date.get():
            return
        date_key = next((c["key"] for c in self.columns if "date" in c["key"].lower()), None)
        if not date_key:
            return
        fmt_map = {
            "YYYY-MM-DD": "%Y-%m-%d", "DD-MM-YYYY": "%d-%m-%Y",
            "DD/MM/YYYY": "%d/%m/%Y", "MM-DD-YYYY": "%m-%d-%Y",
            "MM/DD/YYYY": "%m/%d/%Y", "DD.MM.YYYY": "%d.%m.%Y",
            "DD.MM.YY":   "%d.%m.%y", "DD-MM-YY":   "%d-%m-%y",
            "DD/MM/YY":   "%d/%m/%y",
        }
        from datetime import datetime
        chosen = getattr(self, "var_date_fmt", None)
        fmt = fmt_map.get(chosen.get() if chosen else "YYYY-MM-DD", "%Y-%m-%d")
        def _key(e):
            v = e.get(date_key, "") or ""
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                return datetime.min
        self.session_entries.sort(key=_key)

    def _model_col_key(self, columns=None):
        for c in (columns or self.columns):
            if any(x in c["key"].lower() for x in ("model", "product")):
                return c["key"]
        return None

    def _show_sess_model_popup(self):
        model_key = self._model_col_key()
        if not model_key:
            return
        counts = {}
        for e in self.session_entries:
            v = e.get(model_key, "") or ""
            if v:
                counts[v] = counts.get(v, 0) + 1
        menu = tk.Menu(self.root, tearoff=0)
        total = len(self.session_entries)
        cur = self.sv_sess_model_filter.get().rsplit(" (", 1)[0]
        menu.add_command(label=f"All  ({total})",
                         command=lambda: self.sv_sess_model_filter.set("All"))
        menu.add_separator()
        for m, n in sorted(counts.items()):
            label = f"{'● ' if m == cur else '  '}{m}  ({n})"
            menu.add_command(label=label,
                             command=lambda v=f"{m} ({n})": self.sv_sess_model_filter.set(v))
        menu.post(self.root.winfo_pointerx(), self.root.winfo_pointery())

    def _show_hist_model_popup(self):
        model_key = self._model_col_key(self._hist_effective_columns)
        if not model_key:
            return
        counts = {}
        for e in getattr(self, "_hist_raw_results", []):
            v = e.get(model_key, "") or ""
            if v:
                counts[v] = counts.get(v, 0) + 1
        menu = tk.Menu(self.root, tearoff=0)
        total = sum(counts.values())
        cur = self.sv_hist_model_filter.get().rsplit(" (", 1)[0]
        menu.add_command(label=f"All  ({total})",
                         command=lambda: self.sv_hist_model_filter.set("All"))
        menu.add_separator()
        for m, n in sorted(counts.items()):
            label = f"{'● ' if m == cur else '  '}{m}  ({n})"
            menu.add_command(label=label,
                             command=lambda v=f"{m} ({n})": self.sv_hist_model_filter.set(v))
        menu.post(self.root.winfo_pointerx(), self.root.winfo_pointery())

    def _refresh_tree(self):
        self._sort_by_date()
        self.tree.delete(*self.tree.get_children())

        model_key = self._model_col_key()

        # Determine active model filter
        sel_model = None
        raw = self.sv_sess_model_filter.get()
        if model_key and raw and raw != "All":
            sel_model = raw.rsplit(" (", 1)[0]

        row = 0
        total = len(self.session_entries)
        for i, entry in enumerate(self.session_entries):
            if sel_model and entry.get(model_key, "") != sel_model:
                continue
            tag       = "even" if (row + 1) % 2 == 0 else "odd"
            tree_vals = [entry.get(c["key"], "") for c in self.columns] + ["\u270f", "\U0001f5d1"]
            self.tree.insert("", "end", iid=str(i), values=tree_vals, tags=(tag,))
            row += 1

        # Update model heading to show active filter state
        if model_key:
            col_def = next((c for c in self.columns if c["key"] == model_key), None)
            disp = col_def["display"] if col_def else model_key
            label = f"{disp} [{sel_model}] ▼" if sel_model else f"{disp} ▼"
            try:
                self.tree.heading(model_key, text=label)
            except Exception:
                pass

    # ── icon-column click handling ────────────────────────────────────────────

    def _on_tree_click(self, event):
        col  = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        if not item:
            return
        n         = len(self.columns)
        edit_col  = f"#{n + 1}"
        del_col   = f"#{n + 2}"
        if col not in (edit_col, del_col):
            return
        if self._action_pending:
            return
        self._action_pending = True
        self.root.after(300, self._clear_action_pending)
        idx = int(item)
        if col == edit_col:
            self._load_entry_to_form(idx)
        else:
            self._delete_entry(idx)

    def _clear_action_pending(self):
        self._action_pending = False

    def _load_entry_to_form(self, idx):
        entry = self.session_entries.pop(idx)
        if idx < self.session_saved_count:
            self.session_saved_count = max(0, self.session_saved_count - 1)
        for col_def in self.columns:
            key = col_def["key"]
            if key in self.col_vars:
                self.col_vars[key].set(entry.get(key, ""))
        self._refresh_tree()
        self.lbl_counter.config(
            text=f"Entries this session:  {len(self.session_entries)}")
        self.notebook.select(0)
        self._focus_key_widget()

    def _delete_entry(self, idx):
        entry   = self.session_entries[idx]
        req_key = next((c["key"] for c in self.columns if c.get("required")), None)
        disp    = entry.get(req_key, "this entry") if req_key else "this entry"
        if messagebox.askyesno("Confirm Delete", f"Delete entry for '{disp}'?"):
            self.session_entries.pop(idx)
            if idx < self.session_saved_count:
                self.session_saved_count = max(0, self.session_saved_count - 1)
            self._refresh_tree()
            self.lbl_counter.config(
                text=f"Entries this session:  {len(self.session_entries)}")

    def _undo_last_entry(self):
        if not self.session_entries:
            messagebox.showinfo("Nothing to Undo",
                                "No entries in the current session.")
            return
        self.session_entries.pop()
        if self.session_saved_count > len(self.session_entries):
            self.session_saved_count = len(self.session_entries)
        self._refresh_tree()
        self.lbl_counter.config(
            text=f"Entries this session:  {len(self.session_entries)}")

    # ── find & replace ────────────────────────────────────────────────────────

    def _show_find_replace(self):
        if not self.session_entries:
            messagebox.showinfo("No Entries",
                                "No entries in the current session to search.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Find & Replace")
        dlg.geometry("420x270")
        dlg.resizable(False, False)
        dlg.configure(bg=self.BG)
        dlg.transient(self.root)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Find & Replace  \u2014  Current Session",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=14, pady=10)

        frm = tk.Frame(dlg, bg=self.CARD_BG, padx=16, pady=12)
        frm.pack(fill="both", expand=True)

        field_labels = ["All fields"] + [c["display"] for c in self.columns]
        field_keys   = [None]         + [c["key"]     for c in self.columns]

        sv_field = tk.StringVar(value="All fields")
        sv_find  = tk.StringVar()
        sv_repl  = tk.StringVar()
        sv_case  = tk.BooleanVar(value=False)

        def _row(r, label, make_widget):
            tk.Label(frm, text=label, bg=self.CARD_BG, fg="#555555",
                     font=("Segoe UI", 9)
                     ).grid(row=r, column=0, sticky="w", padx=(0, 10), pady=(4, 0))
            w = make_widget()
            w.grid(row=r, column=1, sticky="ew", pady=(4, 0))
            return w

        _row(0, "Search in:",
             lambda: ttk.Combobox(frm, textvariable=sv_field,
                                  values=field_labels, state="readonly", width=24))
        en_find = _row(1, "Find:",
                       lambda: ttk.Entry(frm, textvariable=sv_find, width=28))
        _row(2, "Replace with:",
             lambda: ttk.Entry(frm, textvariable=sv_repl, width=28))
        ttk.Checkbutton(frm, text="Case sensitive", variable=sv_case
                        ).grid(row=3, column=1, sticky="w", pady=(6, 0))
        frm.columnconfigure(1, weight=1)

        lbl_result = tk.Label(dlg, text="", bg=self.BG, fg=self.DARK_BLUE,
                              font=("Segoe UI", 9, "bold"))
        lbl_result.pack(pady=2)

        def _replace_all():
            import re
            find_txt = sv_find.get()
            if not find_txt:
                messagebox.showwarning("Empty Search", "Enter text to find.", parent=dlg)
                return
            repl_txt = sv_repl.get()
            fidx  = field_labels.index(sv_field.get())
            fkey  = field_keys[fidx]
            flags = 0 if sv_case.get() else re.IGNORECASE
            keys  = [c["key"] for c in self.columns] if fkey is None else [fkey]
            count = 0
            for entry in self.session_entries:
                for k in keys:
                    if k not in entry:
                        continue
                    new_val = re.sub(re.escape(find_txt), repl_txt, entry[k], flags=flags)
                    if new_val != entry[k]:
                        entry[k] = new_val.upper() if k == "pcba_no" else new_val
                        count += 1
            self._refresh_tree()
            lbl_result.config(
                text=f"{count} field{'s' if count != 1 else ''} updated.")

        btn_bar = tk.Frame(dlg, bg=self.BG, padx=14, pady=8)
        btn_bar.pack(fill="x")
        ttk.Button(btn_bar, text="Replace All", style="Dark.TButton",
                   command=_replace_all
                   ).pack(side="left", expand=True, fill="x", padx=(0, 6), ipady=5)
        ttk.Button(btn_bar, text="Close", style="Muted.TButton",
                   command=dlg.destroy
                   ).pack(side="left", expand=True, fill="x", ipady=5)
        en_find.focus_set()

    # ── persistence & export ──────────────────────────────────────────────────

    def save_session(self):
        unsaved = self.session_entries[self.session_saved_count:]
        if not unsaved:
            messagebox.showinfo("Up to Date",
                                "All entries are already saved to the database.")
            return
        # Build label from date/product columns if they exist
        date_key = next((c["key"] for c in self.columns if "date"    in c["key"].lower()), None)
        prod_key = next((c["key"] for c in self.columns if "product" in c["key"].lower()), None)

        if date_key:
            dates = sorted(set(e.get(date_key, "") for e in self.session_entries if e.get(date_key)))
            date_part = dates[0] + (f" \u2013 {dates[-1]}" if len(dates) > 1 else "") if dates else ""
        else:
            date_part = ""

        if prod_key:
            products = sorted(set(e.get(prod_key, "") for e in self.session_entries if e.get(prod_key)))
            prod_part = ", ".join(products[:3])
        else:
            prod_part = ""

        label = " | ".join(filter(None, [prod_part, date_part])) or \
                f"Session {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        save_entries(unsaved, self.session_id, label)
        self.session_saved_count = len(self.session_entries)
        n = len(unsaved)
        messagebox.showinfo("Saved",
                            f"{n} entr{'y' if n == 1 else 'ies'} saved.\nSession: {label}")

    def export_session(self):
        if not self.session_entries:
            messagebox.showinfo("No Entries",
                                "Add at least one entry before exporting.")
            return
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=f"SMT_Rework_{date.today().strftime('%Y%m%d')}.xlsx",
            title="Save Excel Export")
        if not filepath:
            return
        try:
            export_to_excel(self.session_entries, filepath, self.columns)
            self.save_session()
            messagebox.showinfo("Exported", f"Excel file saved:\n{filepath}")
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

    def clear_session(self):
        """Discard the current session, prompting to save unsaved entries first."""
        if not self.session_entries:
            messagebox.showinfo("Nothing to Clear", "The current session is already empty.")
            return

        unsaved = self.session_entries[self.session_saved_count:]
        if unsaved:
            choice = messagebox.askyesnocancel(
                "Unsaved Entries",
                f"You have {len(unsaved)} unsaved entr{'y' if len(unsaved) == 1 else 'ies'}.\n\n"
                "Save before clearing?",
            )
            if choice is None:     # Cancel
                return
            if choice:             # Yes — save first
                self.save_session()

        if not messagebox.askyesno("Clear Session",
                                   "Clear all entries from the current session?\n"
                                   "This cannot be undone."):
            return

        self.session_entries.clear()
        self.session_id          = str(uuid.uuid4())
        self.session_saved_count = 0
        self._refresh_tree()
        self.lbl_session_info.config(text=f"Session: {self.session_id[:8].upper()}")
        for key, var in self.col_vars.items():
            default = next((c["default"] for c in self.columns if c["key"] == key), "")
            var.set(default)
        self._focus_key_widget()

    # ── import Excel ──────────────────────────────────────────────────────────

    def import_excel(self, filepath=None):
        """Import rows from an Excel file; auto-detect columns from its headers."""
        if not filepath:
            filepath = filedialog.askopenfilename(
                parent=self.root,
                filetypes=[("Excel workbook", "*.xlsx *.xls"), ("All files", "*.*")],
                title="Import Excel File")
        if not filepath:
            return

        # Ask whether to clear the current session before importing
        if self.session_entries:
            ans = messagebox.askyesnocancel(
                "Clear Current Session?",
                f"You have {len(self.session_entries)} entr"
                f"{'y' if len(self.session_entries) == 1 else 'ies'} in the current session.\n\n"
                "Yes  — clear session, then import\n"
                "No   — append imported rows to current session\n"
                "Cancel — abort import")
            if ans is None:   # Cancel
                return
            if ans:           # Yes — clear
                self.session_entries.clear()
                self.session_saved_count = 0

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            # Prefer the "Rework Register" sheet; fall back to the first
            # regular worksheet (wb.worksheets excludes chart-only sheets,
            # so this avoids landing on "Chart Data" or an embedded chart sheet).
            if "Rework Register" in wb.sheetnames:
                ws = wb["Rework Register"]
            elif wb.worksheets:
                ws = wb.worksheets[0]
            else:
                ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                messagebox.showwarning("Empty File",
                                       "The selected file contains no data.")
                return

            # ── Step 1: read headers ──────────────────────────────────────────
            raw_hdrs = [str(h).strip() if h is not None else "" for h in rows[0]]
            # Drop trailing empty headers
            while raw_hdrs and not raw_hdrs[-1]:
                raw_hdrs.pop()

            if not raw_hdrs:
                messagebox.showerror("Import Failed",
                                     "No column headers found in the first row.")
                return

            # ── Step 2: map Excel headers to current column keys ──────────────
            def _find_key(hdr):
                hdr_up = hdr.upper().strip()
                # Exact display name match
                for c in self.columns:
                    if c["display"].upper() == hdr_up:
                        return c["key"]
                # Exact key match
                for c in self.columns:
                    if c["key"].upper() == hdr_up:
                        return c["key"]
                # Built-in aliases for backwards-compat
                _aliases = {
                    "DATE": "date",
                    "PRODUCT": "product",
                    "PCBA NO.": "pcba_no", "PCBA NO": "pcba_no",
                    "PCBA NUMBER": "pcba_no", "PCBA": "pcba_no",
                    "PROBLEM REPORTED": "problem", "PROBLEM": "problem",
                    "FAULT": "fault", "FAULT TYPE": "fault", "DEFECT": "fault",
                    "STAGE OF FAULT OBSERVATION": "stage", "STAGE": "stage",
                    "ACTION TAKEN": "action", "ACTION": "action",
                }
                return _aliases.get(hdr_up)

            col_map      = {}   # excel_col_idx → existing key
            unmatched    = []   # (excel_col_idx, header_str) not matched

            for i, hdr in enumerate(raw_hdrs):
                key = _find_key(hdr)
                if key:
                    col_map[i] = key
                else:
                    unmatched.append((i, hdr))

            # ── Step 3: decide whether to sync columns ────────────────────────
            excel_displays  = [h.upper() for h in raw_hdrs]
            current_displays = [c["display"].upper() for c in self.columns]
            columns_match   = (excel_displays == current_displays)

            if not columns_match:
                answer = self._ask_column_sync_dialog(raw_hdrs, unmatched)
                if answer == "cancel":
                    return
                if answer == "sync":
                    new_columns = []
                    for i, hdr in enumerate(raw_hdrs):
                        if i in col_map:
                            # Re-use existing col def but update display to match Excel
                            existing = next((c for c in self.columns if c["key"] == col_map[i]), None)
                            if existing:
                                new_col = dict(existing)
                                new_col["display"] = hdr
                                new_columns.append(new_col)
                            else:
                                new_columns.append({
                                    "key": col_map[i], "display": hdr,
                                    "type": "entry", "options": [],
                                    "required": False, "default": "",
                                })
                        else:
                            # Brand-new column
                            raw_key = hdr.lower().strip()
                            raw_key = "".join(
                                ch if ch.isalnum() or ch == "_" else "_" for ch in raw_key.replace(" ", "_").replace(".", ""))
                            # Ensure uniqueness
                            existing_keys = {c["key"] for c in new_columns}
                            base, suffix = raw_key, 2
                            while raw_key in existing_keys:
                                raw_key = f"{base}_{suffix}"
                                suffix += 1
                            col_map[i] = raw_key
                            new_columns.append({
                                "key": raw_key, "display": hdr,
                                "type": "entry", "options": [],
                                "required": False, "default": "",
                            })
                    self.columns = new_columns
                    save_columns_cfg(self.columns)
                    self._apply_column_changes()
                    # Rebuild col_map directly from new_columns order — new_columns[i]
                    # was built to correspond exactly to raw_hdrs[i], so this is safe
                    # and guarantees the keys match the updated self.columns.
                    col_map = {i: new_columns[i]["key"] for i in range(len(new_columns))}
                # if "import_as_is": proceed with col_map as-is (partial mapping)

            # ── Step 4: import rows ───────────────────────────────────────────
            def _cell(row, col_idx):
                if col_idx >= len(row):
                    return ""
                val = row[col_idx]
                if val is None:
                    return ""
                if hasattr(val, "strftime"):
                    return val.strftime("%Y-%m-%d")
                return str(val).strip()

            def _is_blank_row(row):
                """True if every cell is None or whitespace — catches blank separator rows."""
                return all(
                    c is None or (isinstance(c, str) and not c.strip())
                    for c in row
                )

            req_key = next((c["key"] for c in self.columns if c.get("required")), None)
            imported = skipped = 0

            for row in rows[1:]:
                if _is_blank_row(row):
                    continue
                entry = {}
                for excel_idx, key in col_map.items():
                    val = _cell(row, excel_idx)
                    if key == "pcba_no":
                        val = val.upper()
                    entry[key] = val

                # Skip rows missing the required field
                if req_key and not entry.get(req_key, "").strip():
                    skipped += 1
                    continue

                self.session_entries.append(entry)
                imported += 1

            self._refresh_tree()
            self.lbl_counter.config(
                text=f"Entries this session:  {len(self.session_entries)}")

            msg = f"{imported} entr{'y' if imported == 1 else 'ies'} imported."
            if skipped:
                msg += f"\n{skipped} row{'s' if skipped > 1 else ''} skipped (no required field)."

            # Warn about app columns that had no matching Excel column
            mapped_keys       = set(col_map.values())
            unmapped_app_cols = [c["display"] for c in self.columns
                                 if c["key"] not in mapped_keys]
            if unmapped_app_cols:
                msg += (f"\n\n⚠ The following columns were not found in the "
                        f"Excel file — those cells will be blank:\n  "
                        + ", ".join(unmapped_app_cols))

            icon = "warning" if unmapped_app_cols else "info"
            messagebox.showinfo("Import Complete", msg) if icon == "info" \
                else messagebox.showwarning("Import Complete", msg)

        except Exception as exc:
            messagebox.showerror("Import Failed", str(exc))

    def _ask_column_sync_dialog(self, excel_hdrs, unmatched):
        """Show dialog asking whether to sync app columns with the Excel file's headers.
        Returns 'sync', 'import_as_is', or 'cancel'.
        """
        result = tk.StringVar(value="cancel")

        dlg = tk.Toplevel(self.root)
        dlg.title("Column Mismatch")
        dlg.geometry("500x340")
        dlg.resizable(False, True)
        dlg.configure(bg=self.BG)
        dlg.transient(self.root)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Excel Column Mismatch",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=14, pady=10)

        body = tk.Frame(dlg, bg=self.CARD_BG, padx=16, pady=12)
        body.pack(fill="both", expand=True)

        tk.Label(body, text="This Excel file has columns:",
                 bg=self.CARD_BG, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        tk.Label(body, text=", ".join(excel_hdrs),
                 bg=self.CARD_BG, fg="#333333",
                 font=("Segoe UI", 9), wraplength=460, justify="left"
                 ).pack(anchor="w", pady=(2, 6))

        if unmatched:
            new_names = [h for _, h in unmatched]
            tk.Label(body,
                     text=f"New columns not in current setup: {', '.join(new_names)}",
                     bg=self.CARD_BG, fg="#B03030",
                     font=("Segoe UI", 9), wraplength=460, justify="left"
                     ).pack(anchor="w", pady=(0, 6))

        tk.Label(body,
                 text="Current app columns:  " + ", ".join(c["display"] for c in self.columns),
                 bg=self.CARD_BG, fg="#555555",
                 font=("Segoe UI", 9), wraplength=460, justify="left"
                 ).pack(anchor="w", pady=(0, 8))

        tk.Label(body, text="How would you like to import?",
                 bg=self.CARD_BG, font=("Segoe UI", 9)).pack(anchor="w")

        def _choose(val):
            result.set(val)
            dlg.destroy()

        btn_frame = tk.Frame(dlg, bg=self.BG, padx=14, pady=10)
        btn_frame.pack(fill="x")

        ttk.Button(btn_frame,
                   text="Sync Columns with Excel  (adopt Excel headers)",
                   style="Dark.TButton",
                   command=lambda: _choose("sync")
                   ).pack(fill="x", ipady=5, pady=(0, 4))
        ttk.Button(btn_frame,
                   text="Keep Current Columns  (import with best-effort mapping)",
                   style="Muted.TButton",
                   command=lambda: _choose("import_as_is")
                   ).pack(fill="x", ipady=5, pady=(0, 4))
        ttk.Button(btn_frame,
                   text="Cancel",
                   style="Danger.TButton",
                   command=lambda: _choose("cancel")
                   ).pack(fill="x", ipady=5)

        dlg.protocol("WM_DELETE_WINDOW", lambda: _choose("cancel"))
        dlg.wait_window()
        return result.get()

    def _sync_columns_to_entries(self, entries):
        """Adjust self.columns to match the keys present in entries.

        Preserves existing column defs for matching keys; creates minimal defs
        for new keys; removes columns whose keys are absent from every entry.
        Does NOT save to disk — the user's columns.json is unchanged.
        """
        if not entries:
            return
        # Collect ordered unique keys as they appear across entries
        seen = {}
        for e in entries:
            for k in e:
                if k not in seen:
                    seen[k] = True
        session_keys = list(seen.keys())
        current_by_key = {c["key"]: c for c in self.columns}
        new_columns = []
        for k in session_keys:
            if k in current_by_key:
                new_columns.append(current_by_key[k])
            else:
                new_columns.append({
                    "key": k, "display": k.replace("_", " ").title(),
                    "type": "entry", "options": [],
                    "required": False, "default": "",
                })
        if [c["key"] for c in new_columns] == [c["key"] for c in self.columns]:
            return  # no change needed
        self.columns = new_columns
        self._apply_column_changes()

    def _apply_column_changes(self):
        """Rebuild form fields and both treeviews after columns change."""
        self._populate_form_fields()
        self._create_treeview()
        self._create_hist_treeview()
        self._rebuild_hist_filter_row2()

    # ── Manage Columns dialog ─────────────────────────────────────────────────

    def _show_manage_columns(self):
        import copy
        working = copy.deepcopy(self.columns)
        # Tag each column with its original key so renames can be tracked
        for c in working:
            c["_orig_key"] = c["key"]

        dlg = tk.Toplevel(self.root)
        dlg.title("Manage Columns")
        dlg.geometry("600x520")
        dlg.resizable(True, True)
        dlg.configure(bg=self.BG)
        dlg.transient(self.root)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Manage Columns",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=14, pady=10)

        main = tk.Frame(dlg, bg=self.BG)
        main.pack(fill="both", expand=True, padx=10, pady=8)

        # ── Left: column list ─────────────────────────────────────────────────
        list_pane = tk.Frame(main, bg=self.CARD_BG, width=190)
        list_pane.pack(side="left", fill="y", padx=(0, 6))
        list_pane.pack_propagate(False)

        tk.Label(list_pane, text="Columns (in order)",
                 bg=self.CARD_BG, fg=self.DARK_BLUE,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=8, pady=(6, 2))

        lb = tk.Listbox(list_pane, font=("Segoe UI", 9),
                        bg=self.CARD_BG, selectbackground="#BDD7EE",
                        relief="flat", bd=0, activestyle="none")
        lb.pack(fill="both", expand=True, padx=4, pady=2)

        current_idx = [None]

        def refresh_list(sel=None):
            lb.delete(0, "end")
            for c in working:
                req = " ★" if c.get("required") else ""
                lb.insert("end", c["display"] + req)
            if sel is not None and sel < len(working):
                lb.selection_set(sel)
                lb.see(sel)
                current_idx[0] = sel

        refresh_list(0)

        # List action buttons
        lbtn = tk.Frame(list_pane, bg=self.CARD_BG)
        lbtn.pack(fill="x", padx=4, pady=(0, 4))

        def add_col():
            n = len(working) + 1
            working.append({
                "key": f"col_{n}", "display": f"NEW COLUMN {n}",
                "type": "entry", "options": [], "required": False, "default": "",
                "_orig_key": f"col_{n}",
            })
            refresh_list(len(working) - 1)
            load_editor(len(working) - 1)

        def del_col():
            sel = lb.curselection()
            if not sel:
                return
            idx = sel[0]
            col = working[idx]
            if not messagebox.askyesno(
                    "Delete Column",
                    f"Remove column '{col['display']}' from the form?\n\n"
                    "Data stored under its key will remain in the database.",
                    parent=dlg):
                return
            working.pop(idx)
            current_idx[0] = None
            new_sel = min(idx, len(working) - 1) if working else None
            refresh_list(new_sel)
            if new_sel is not None:
                load_editor(new_sel)

        def move_up():
            sel = lb.curselection()
            if not sel or sel[0] == 0:
                return
            i = sel[0]
            working[i - 1], working[i] = working[i], working[i - 1]
            refresh_list(i - 1)

        def move_down():
            sel = lb.curselection()
            if not sel or sel[0] >= len(working) - 1:
                return
            i = sel[0]
            working[i + 1], working[i] = working[i], working[i + 1]
            refresh_list(i + 1)

        row1 = tk.Frame(lbtn, bg=self.CARD_BG)
        row1.pack(fill="x", pady=(0, 2))
        row2 = tk.Frame(lbtn, bg=self.CARD_BG)
        row2.pack(fill="x")
        for txt, fn, sty, parent in [
            ("+ Add",  add_col,   "Dark.TButton",   row1),
            ("Delete", del_col,   "Danger.TButton", row1),
            ("\u2191 Up",   move_up,   "Muted.TButton",  row2),
            ("\u2193 Down", move_down, "Muted.TButton",  row2),
        ]:
            ttk.Button(parent, text=txt, style=sty, command=fn
                       ).pack(side="left", padx=(0, 2), ipady=3, ipadx=6, expand=True, fill="x")

        # ── Right: column editor ──────────────────────────────────────────────
        edit_pane = tk.Frame(main, bg=self.CARD_BG)
        edit_pane.pack(side="left", fill="both", expand=True)

        tk.Label(edit_pane, text="Column Properties",
                 bg=self.CARD_BG, fg=self.DARK_BLUE,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=8, pady=(6, 2))

        props = tk.Frame(edit_pane, bg=self.CARD_BG, padx=12, pady=8)
        props.pack(fill="both", expand=True)

        sv_display  = tk.StringVar()
        sv_key      = tk.StringVar()
        sv_type     = tk.StringVar(value="entry")
        sv_required = tk.BooleanVar(value=False)
        sv_default  = tk.StringVar()

        def _plbl(text, row):
            tk.Label(props, text=text, bg=self.CARD_BG,
                     fg="#555555", font=("Segoe UI", 9)
                     ).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=(5, 0))

        _plbl("Display Name:",  0); ttk.Entry(props, textvariable=sv_display, width=22).grid(row=0, column=1, sticky="ew", pady=(5, 0))
        _plbl("Key (internal):", 1); ttk.Entry(props, textvariable=sv_key,     width=22).grid(row=1, column=1, sticky="ew", pady=(5, 0))
        _plbl("Type:",           2)
        ttk.Combobox(props, textvariable=sv_type,
                     values=["entry", "combo"], state="readonly", width=12
                     ).grid(row=2, column=1, sticky="w", pady=(5, 0))
        ttk.Checkbutton(props, text="Required field", variable=sv_required
                        ).grid(row=3, column=1, sticky="w", pady=(6, 0))
        _plbl("Default value:", 4); ttk.Entry(props, textvariable=sv_default, width=22).grid(row=4, column=1, sticky="ew", pady=(5, 0))

        _plbl("Options\n(comma-separated,\nfor combo):", 5)
        txt_opts = tk.Text(props, height=4, width=22, font=("Segoe UI", 9),
                           relief="solid", bd=1)
        txt_opts.grid(row=5, column=1, sticky="ew", pady=(5, 0))

        props.columnconfigure(1, weight=1)

        def load_editor(idx):
            if idx is None or idx >= len(working):
                return
            c = working[idx]
            sv_display.set(c.get("display", ""))
            sv_key.set(c.get("key", ""))
            sv_type.set(c.get("type", "entry"))
            sv_required.set(c.get("required", False))
            sv_default.set(c.get("default", ""))
            txt_opts.delete("1.0", "end")
            txt_opts.insert("1.0", ", ".join(c.get("options", [])))
            current_idx[0] = idx

        def save_editor():
            idx = current_idx[0]
            if idx is None or idx >= len(working):
                messagebox.showwarning("No Column Selected",
                                       "Select a column from the list first.", parent=dlg)
                return
            display = sv_display.get().strip()
            key     = sv_key.get().strip()
            if not display or not key:
                messagebox.showwarning("Invalid", "Display name and key cannot be empty.",
                                       parent=dlg)
                return
            for i, c in enumerate(working):
                if i != idx and c["key"] == key:
                    messagebox.showwarning("Duplicate Key",
                                          f"Key '{key}' is already used by column '{c['display']}'.",
                                          parent=dlg)
                    return
            opts_raw = txt_opts.get("1.0", "end-1c").strip()
            options  = [o.strip() for o in opts_raw.split(",") if o.strip()]
            working[idx].update({
                "display":  display,
                "key":      key,
                "type":     sv_type.get(),
                "required": sv_required.get(),
                "default":  sv_default.get().strip(),
                "options":  options,
            })
            refresh_list(idx)
            messagebox.showinfo("Saved", f"Column '{display}' updated in the list.\n"
                                         "Click 'Save & Apply' to apply to the app.", parent=dlg)

        def on_select(event=None):
            sel = lb.curselection()
            if sel:
                load_editor(sel[0])

        lb.bind("<<ListboxSelect>>", on_select)

        ttk.Button(props, text="Update Column in List", style="Dark.TButton",
                   command=save_editor
                   ).grid(row=6, column=0, columnspan=2, sticky="ew", pady=(10, 0), ipady=4)

        if working:
            load_editor(0)

        # ── Bottom bar ────────────────────────────────────────────────────────
        def _apply():
            if not working:
                messagebox.showwarning("No Columns",
                                       "At least one column is required.", parent=dlg)
                return

            # Silently flush any unsaved edits from the editor panel into working
            _flush_idx = current_idx[0]
            if _flush_idx is not None and _flush_idx < len(working):
                _disp = sv_display.get().strip()
                _key  = sv_key.get().strip()
                if _disp and _key:
                    _dup = any(i != _flush_idx and working[i]["key"] == _key
                               for i in range(len(working)))
                    if not _dup:
                        _opts = [o.strip() for o in
                                 txt_opts.get("1.0", "end-1c").strip().split(",") if o.strip()]
                        working[_flush_idx].update({
                            "display":  _disp, "key": _key,
                            "type":     sv_type.get(), "required": sv_required.get(),
                            "default":  sv_default.get().strip(), "options": _opts,
                        })

            # Detect brand-new columns that have a non-empty default value
            old_keys = {c["key"] for c in self.columns}
            new_with_default = [
                c for c in working
                if c["key"] not in old_keys and c.get("default", "").strip()
            ]

            # If there are existing session rows, ask whether to backfill
            if new_with_default and self.session_entries:
                lines = "\n".join(
                    f"  • {c['display']}  →  \"{c['default']}\""
                    for c in new_with_default
                )
                ans = messagebox.askyesnocancel(
                    "Apply Default to Existing Rows?",
                    f"The following new column(s) have a default value:\n\n"
                    f"{lines}\n\n"
                    f"Apply that default to all {len(self.session_entries)} "
                    f"existing row(s)?\n\n"
                    "Yes — fill existing rows\n"
                    "No  — leave existing rows blank for these columns\n"
                    "Cancel — go back",
                    parent=dlg)
                if ans is None:   # Cancel — abort
                    return
                if ans:           # Yes — backfill
                    for entry in self.session_entries:
                        for col in new_with_default:
                            entry.setdefault(col["key"], col["default"])

            # Build rename map: orig_key → new_key for any renamed columns
            rename_map = {}
            for c in working:
                orig = c.pop("_orig_key", None)
                if orig and orig != c["key"]:
                    rename_map[orig] = c["key"]
            # Migrate session entries so renamed keys don't lose data
            if rename_map and self.session_entries:
                for entry in self.session_entries:
                    for old_k, new_k in rename_map.items():
                        if old_k in entry:
                            entry[new_k] = entry.pop(old_k)
            self.columns = list(working)
            save_columns_cfg(self.columns)
            self._apply_column_changes()
            dlg.destroy()
            messagebox.showinfo("Columns Applied",
                                f"{len(self.columns)} column(s) configured and applied.")

        def _reset():
            if messagebox.askyesno("Reset to Defaults",
                                   "Reset columns to the factory defaults?\n"
                                   "Your current working changes will be lost.", parent=dlg):
                working.clear()
                working.extend([dict(c) for c in DEFAULT_COLUMNS])
                refresh_list(0)
                load_editor(0)

        btn_bar = tk.Frame(dlg, bg=self.BG, padx=14, pady=8)
        btn_bar.pack(fill="x")
        ttk.Button(btn_bar, text="Save & Apply",      style="Green.TButton",
                   command=_apply  ).pack(side="left", expand=True, fill="x", padx=(0, 4), ipady=5)
        ttk.Button(btn_bar, text="Reset to Defaults", style="Muted.TButton",
                   command=_reset  ).pack(side="left", padx=(0, 4), ipady=5, ipadx=8)
        ttk.Button(btn_bar, text="Cancel",            style="Muted.TButton",
                   command=dlg.destroy).pack(side="left", expand=True, fill="x", ipady=5)

    # ── History & Search tab ──────────────────────────────────────────────────

    def _build_history_tab(self, parent):
        title_bar = tk.Frame(parent, bg=self.DARK_BLUE, height=40)
        title_bar.pack(fill="x")
        title_bar.pack_propagate(False)
        tk.Label(title_bar, text="History & Search",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=10)

        self._hist_fbar = tk.Frame(parent, bg="#E4EDF5")
        self._hist_fbar.pack(fill="x")

        row1 = tk.Frame(self._hist_fbar, bg="#E4EDF5")
        row1.pack(fill="x", padx=6, pady=(6, 2))

        def _flbl(parent_row, text):
            tk.Label(parent_row, text=text, bg="#E4EDF5",
                     fg="#333333", font=("Segoe UI", 9)
                     ).pack(side="left", padx=(8, 3))

        _flbl(row1, "From:")
        self.sv_hist_from = tk.StringVar()
        ttk.Entry(row1, textvariable=self.sv_hist_from,
                  width=12, font=("Segoe UI", 9)).pack(side="left", padx=(0, 6), ipady=2)

        _flbl(row1, "To:")
        self.sv_hist_to = tk.StringVar()
        ttk.Entry(row1, textvariable=self.sv_hist_to,
                  width=12, font=("Segoe UI", 9)).pack(side="left", padx=(0, 6), ipady=2)

        _flbl(row1, "Search:")
        self.sv_hist_pcba = tk.StringVar()
        ttk.Entry(row1, textvariable=self.sv_hist_pcba,
                  width=18, font=("Segoe UI", 9)).pack(side="left", padx=(0, 6), ipady=2)

        _flbl(row1, "Model:")
        self.sv_hist_model_filter = tk.StringVar(value="All")
        self.cmb_hist_model = ttk.Combobox(row1, textvariable=self.sv_hist_model_filter,
                                           state="readonly", width=16)
        self.cmb_hist_model.pack(side="left", padx=(0, 12), ipady=2)
        self.sv_hist_model_filter.trace_add("write", lambda *_: self._apply_hist_model_filter())

        ttk.Button(row1, text="Search",   style="Dark.TButton",
                   command=self._hist_search).pack(side="left", padx=(0, 6), ipady=3, ipadx=10)
        ttk.Button(row1, text="Show All", style="Green.TButton",
                   command=self._hist_show_all).pack(side="left", padx=(0, 6), ipady=3, ipadx=8)
        ttk.Button(row1, text="Clear",    style="Muted.TButton",
                   command=self._hist_clear).pack(side="left", ipady=3, ipadx=8)

        # row2: dynamic combo filters — stored as a Frame reference
        self._hist_row2 = tk.Frame(self._hist_fbar, bg="#E4EDF5")
        self._hist_row2.pack(fill="x", padx=6, pady=(0, 6))
        self._rebuild_hist_filter_row2()
        self.root.after(100, self._update_hist_model_options)

        self._build_hist_bottom_bar(parent)

        content = ttk.Frame(parent)
        content.pack(fill="both", expand=True, padx=10, pady=(8, 4))

        sess_panel = tk.Frame(content, bg=self.CARD_BG, width=215)
        sess_panel.pack(side="left", fill="y", padx=(0, 8))
        sess_panel.pack_propagate(False)
        self._build_sessions_panel(sess_panel)

        results_panel = tk.Frame(content, bg=self.CARD_BG)
        results_panel.pack(side="left", fill="both", expand=True)
        self._build_hist_results(results_panel)

    def _rebuild_hist_filter_row2(self):
        """Rebuild the combo-filter row in History based on current columns."""
        for w in self._hist_row2.winfo_children():
            w.destroy()
        self.hist_combo_filters = []

        combo_cols = [c for c in self.columns if c.get("type") == "combo" and c.get("options")][:3]
        for col_def in combo_cols:
            key  = col_def["key"]
            disp = col_def["display"]
            sv   = tk.StringVar(value="All")
            self.hist_combo_filters.append((key, sv))
            tk.Label(self._hist_row2, text=f"{disp}:", bg="#E4EDF5",
                     fg="#333333", font=("Segoe UI", 9)
                     ).pack(side="left", padx=(8, 3))
            ttk.Combobox(self._hist_row2, textvariable=sv,
                         values=["All"] + col_def["options"],
                         state="readonly", width=16,
                         font=("Segoe UI", 9)).pack(side="left", padx=(0, 10), ipady=2)

        tk.Label(self._hist_row2,
                 text="(leave blank / \u2018All\u2019 to skip filter)",
                 bg="#E4EDF5", fg="#909090",
                 font=("Segoe UI", 8, "italic")).pack(side="left", padx=(8, 0))

    def _build_sessions_panel(self, parent):
        hdr = tk.Frame(parent, bg=self.DARK_BLUE, height=36)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Past Sessions",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=8)

        lb_frame = tk.Frame(parent, bg=self.CARD_BG)
        lb_frame.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(lb_frame, orient="vertical")
        vsb.pack(side="right", fill="y")

        self.lb_sessions = tk.Listbox(
            lb_frame,
            yscrollcommand=vsb.set,
            font=("Segoe UI", 9),
            bg=self.CARD_BG, fg="#1A1A1A",
            selectbackground="#BDD7EE", selectforeground="black",
            relief="flat", bd=0, activestyle="none", cursor="hand2")
        vsb.config(command=self.lb_sessions.yview)
        self.lb_sessions.pack(fill="both", expand=True, padx=2, pady=2)
        self.lb_sessions.bind("<<ListboxSelect>>", self._on_session_select)

        self._hist_sessions_data = []
        self._hist_refresh_sessions()

        btn_frame = tk.Frame(parent, bg=self.CARD_BG, padx=6, pady=6)
        btn_frame.pack(fill="x")
        ttk.Button(btn_frame, text="Open as Current Session",
                   style="Dark.TButton",
                   command=self._load_session_as_current
                   ).pack(fill="x", ipady=4, pady=(0, 4))
        ttk.Button(btn_frame, text="Rename Session",
                   style="Muted.TButton",
                   command=self._rename_session
                   ).pack(fill="x", ipady=4, pady=(0, 4))
        ttk.Button(btn_frame, text="Delete Session",
                   style="Danger.TButton",
                   command=self._delete_session
                   ).pack(fill="x", ipady=4)

    def _build_hist_results(self, parent):
        hdr = tk.Frame(parent, bg=self.DARK_BLUE, height=36)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Search Results",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=8)

        self.lbl_row_count = tk.Label(
            parent, text="0 rows  ",
            bg=self.CARD_BG, fg="#909090", font=("Segoe UI", 9))
        self.lbl_row_count.pack(side="bottom", anchor="e", padx=8, pady=4)

        self.hist_tv_frame = tk.Frame(parent, bg=self.CARD_BG)
        self.hist_tv_frame.pack(fill="both", expand=True)

        self._create_hist_treeview()

    def _create_hist_treeview(self, columns=None):
        """Create / recreate the history treeview with the given column layout.
        If *columns* is None the current self.columns is used (e.g. after a
        Manage-Columns change).  When displaying search results that span
        multiple sessions with different schemas the caller passes the union
        of all relevant column definitions.
        """
        if columns is None:
            columns = self.columns
        self._hist_effective_columns = columns   # used by click handler & exports

        for w in self.hist_tv_frame.winfo_children():
            w.destroy()

        data_cols = tuple(c["key"] for c in columns)
        all_cols  = data_cols + ("session", "edit_col", "del_col")

        self.hist_tree = ttk.Treeview(self.hist_tv_frame, columns=all_cols,
                                      show="headings", selectmode="extended")

        hist_model_key = self._model_col_key(columns)
        for col_def in columns:
            key  = col_def["key"]
            disp = col_def["display"]
            w    = max(65, min(180, len(disp) * 9 + 10))
            if key == hist_model_key:
                sel = getattr(self, "sv_hist_model_filter", None)
                sel_val = sel.get().rsplit(" (", 1)[0] if sel and sel.get() != "All" else None
                hlabel = f"{disp} [{sel_val}] ▼" if sel_val else f"{disp} ▼"
                self.hist_tree.heading(key, text=hlabel,
                                       command=self._show_hist_model_popup)
            else:
                self.hist_tree.heading(key, text=disp)
            self.hist_tree.column(key, width=w, anchor="center", minwidth=50)

        self.hist_tree.heading("session",  text="Session")
        self.hist_tree.column("session",   width=135, anchor="w",      minwidth=50)
        self.hist_tree.heading("edit_col", text="\u270f")
        self.hist_tree.column("edit_col",  width=34,  anchor="center", minwidth=34)
        self.hist_tree.heading("del_col",  text="\U0001f5d1")
        self.hist_tree.column("del_col",   width=34,  anchor="center", minwidth=34)

        self.hist_tree.tag_configure("odd",  background=self.CARD_BG)
        self.hist_tree.tag_configure("even", background=self.ALT_ROW)

        vsb = ttk.Scrollbar(self.hist_tv_frame, orient="vertical",   command=self.hist_tree.yview)
        hsb = ttk.Scrollbar(self.hist_tv_frame, orient="horizontal", command=self.hist_tree.xview)
        self.hist_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right",  fill="y")
        self.hist_tree.pack(fill="both", expand=True)

        self.hist_tree.bind("<Button-1>", self._hist_on_tree_click)

    def _build_hist_bottom_bar(self, parent):
        bar = tk.Frame(parent, bg=self.BG, height=52)
        bar.pack(side="bottom", fill="x", padx=10, pady=(0, 8))
        bar.pack_propagate(False)

        ttk.Button(bar, text="Export Selected Session to Excel",
                   style="Dark.TButton",
                   command=self._hist_export_session
                   ).pack(side="left", pady=10, ipady=5, ipadx=14)

        ttk.Button(bar, text="Export All Results to Excel",
                   style="Green.TButton",
                   command=self._hist_export_all
                   ).pack(side="left", padx=8, pady=10, ipady=5, ipadx=14)

        ttk.Button(bar, text="Export Selected Rows",
                   style="Muted.TButton",
                   command=self._hist_export_selected
                   ).pack(side="left", pady=10, ipady=5, ipadx=10)

        ttk.Button(bar, text="Delete Selected",
                   style="Danger.TButton",
                   command=self._hist_delete_selected
                   ).pack(side="left", padx=8, pady=10, ipady=5, ipadx=10)

    # ── history tab logic ─────────────────────────────────────────────────────

    def _hist_refresh_sessions(self):
        self._hist_sessions_data = get_all_sessions()
        self.lb_sessions.delete(0, "end")
        for sess in self._hist_sessions_data:
            label   = sess["label"]
            display = label if len(label) <= 27 else label[:25] + "\u2026"
            self.lb_sessions.insert("end", display)

    def _on_session_select(self, _event=None):
        sel = self.lb_sessions.curselection()
        if not sel:
            return
        session = self._hist_sessions_data[sel[0]]
        entries = get_session_entries(session["session_id"])
        for e in entries:
            e["session"] = session["label"]
        self._hist_raw_results = list(entries)
        self._update_hist_model_options()
        self._apply_hist_model_filter()

    def _apply_hist_model_filter(self):
        """Re-filter _hist_raw_results by the model combobox and repopulate."""
        if not hasattr(self, "_hist_raw_results"):
            return
        model_key = self._model_col_key()
        raw = self.sv_hist_model_filter.get()
        sel_model = raw.rsplit(" (", 1)[0] if raw and raw != "All" else None
        if sel_model and model_key:
            filtered = [r for r in self._hist_raw_results
                        if r.get(model_key, "") == sel_model]
        else:
            filtered = self._hist_raw_results
        self._populate_hist_tree(filtered)

    def _hist_search(self):
        date_from = self.sv_hist_from.get().strip() or None
        date_to   = self.sv_hist_to.get().strip()   or None
        text_search = self.sv_hist_pcba.get().strip() or None
        results     = search_entries(date_from=date_from, date_to=date_to,
                                     text_search=text_search)

        # Apply dynamic combo filters
        for key, sv in self.hist_combo_filters:
            val = sv.get()
            if val and val != "All":
                results = [r for r in results if r.get(key, "") == val]

        for r in results:
            r["session"] = r.get("session_label", "")
        self.lb_sessions.selection_clear(0, "end")
        self._hist_raw_results = list(results)
        self._update_hist_model_options()
        self._apply_hist_model_filter()

    def _hist_show_all(self):
        results = search_entries()
        for r in results:
            r["session"] = r.get("session_label", "")
        self.lb_sessions.selection_clear(0, "end")
        self._hist_raw_results = list(results)
        self._update_hist_model_options()
        self._apply_hist_model_filter()

    def _update_hist_model_options(self):
        if not hasattr(self, "cmb_hist_model"):
            return
        # Always show all models from DB so user can filter before/after searching
        options = ["All"] + get_all_model_values()
        self.cmb_hist_model["values"] = options
        if self.sv_hist_model_filter.get() not in options:
            self.sv_hist_model_filter.set("All")

    def _hist_clear(self):
        self.sv_hist_from.set("")
        self.sv_hist_to.set("")
        self.sv_hist_pcba.set("")
        for _, sv in self.hist_combo_filters:
            sv.set("All")
        if hasattr(self, "sv_hist_model_filter"):
            self.sv_hist_model_filter.set("All")
        self._hist_raw_results = []
        self.lb_sessions.selection_clear(0, "end")
        self._populate_hist_tree([])

    def _populate_hist_tree(self, entries):
        # ── Compute effective columns from the actual entry keys ───────────────
        # This ensures sessions stored with a different column schema still show
        # their data correctly, regardless of the current self.columns config.
        if entries:
            # Collect all distinct data keys that appear in any entry (order-stable)
            seen_keys = {}   # key → first col_def that matches, or synthetic
            current_col_map = {c["key"]: c for c in self.columns}
            for e in entries:
                for k in e.keys():
                    if k not in _META_KEYS and k not in seen_keys:
                        if k in current_col_map:
                            seen_keys[k] = current_col_map[k]
                        else:
                            # Synthetic column def for keys not in current config
                            seen_keys[k] = {
                                "key":      k,
                                "display":  k.replace("_", " ").title(),
                                "type":     "entry",
                                "options":  [],
                                "required": False,
                                "default":  "",
                            }
            # Prefer the current columns order; then append any extras
            effective_cols = [c for c in self.columns if c["key"] in seen_keys]
            for k, col_def in seen_keys.items():
                if k not in current_col_map:
                    effective_cols.append(col_def)
        else:
            effective_cols = self.columns

        # Always rebuild the treeview — ensures columns are never stale between
        # sessions that have different schemas.
        self._create_hist_treeview(effective_cols)

        self.hist_tree.delete(*self.hist_tree.get_children())
        for i, e in enumerate(entries):
            tag       = "even" if (i + 1) % 2 == 0 else "odd"
            data_vals = [e.get(c["key"], "") for c in self._hist_effective_columns]
            self.hist_tree.insert(
                "", "end", iid=str(i),
                values=data_vals + [
                    e.get("session", e.get("session_label", "")),
                    "\u270f", "\U0001f5d1",
                ],
                tags=(tag,))
        n = len(entries)
        self.lbl_row_count.config(text=f"{n} row{'s' if n != 1 else ''}  ")
        self._hist_current_results = list(entries)

    def _hist_export_session(self):
        sel = self.lb_sessions.curselection()
        if not sel:
            messagebox.showinfo("No Session Selected",
                                "Click a session in the panel on the left first.")
            return
        session  = self._hist_sessions_data[sel[0]]
        entries  = get_session_entries(session["session_id"])
        filename = f"SMT_Session_{session['created_at'][:10]}.xlsx"
        self._do_export(entries, filename)

    def _hist_export_all(self):
        if not self._hist_current_results:
            messagebox.showinfo("No Results",
                                "Run a search or select a session first.")
            return
        filename = f"SMT_Search_{date.today().strftime('%Y%m%d')}.xlsx"
        self._do_export(self._hist_current_results, filename)

    def _do_export(self, entries, default_filename):
        if not entries:
            messagebox.showinfo("No Entries", "Nothing to export.")
            return
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=default_filename,
            title="Save Excel Export")
        if not filepath:
            return
        try:
            export_to_excel(entries, filepath, self.columns)
            messagebox.showinfo("Exported", f"Excel file saved:\n{filepath}")
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

    # ── history row icon actions ──────────────────────────────────────────────

    def _hist_on_tree_click(self, event):
        col  = self.hist_tree.identify_column(event.x)
        item = self.hist_tree.identify_row(event.y)
        if not item:
            return
        n        = len(self._hist_effective_columns)
        edit_col = f"#{n + 2}"   # data(n) + session(1) + edit
        del_col  = f"#{n + 3}"   # data(n) + session(1) + edit + del
        if col not in (edit_col, del_col):
            return   # let Tkinter handle selection normally
        if self._hist_action_pending:
            return "break"
        self._hist_action_pending = True
        self.root.after(300, self._clear_hist_action_pending)
        idx = int(item)
        if col == edit_col:
            self._hist_edit_entry(idx)
        else:
            self._hist_delete_entry(idx)
        return "break"   # preserve multi-selection when clicking icon columns

    def _clear_hist_action_pending(self):
        self._hist_action_pending = False

    def _hist_edit_entry(self, idx):
        entry    = self._hist_current_results[idx]
        entry_id = entry["id"]

        req_col  = next((c for c in self.columns if c.get("required")), None)
        req_key  = req_col["key"] if req_col else None
        req_disp = entry.get(req_key, "Entry") if req_key else "Entry"

        dlg = tk.Toplevel(self.root)
        dlg.title("Edit Entry")
        dlg.geometry("360x440")
        dlg.resizable(True, True)
        dlg.configure(bg=self.BG)
        dlg.transient(self.root)
        dlg.grab_set()

        hdr = tk.Frame(dlg, bg=self.DARK_BLUE, height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"Edit Entry  \u2014  {req_disp}",
                 bg=self.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=14, pady=10)

        # Scrollable form
        outer = tk.Frame(dlg, bg=self.CARD_BG)
        outer.pack(fill="both", expand=True, padx=14, pady=8)
        canvas = tk.Canvas(outer, bg=self.CARD_BG, highlightthickness=0)
        vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        frm = tk.Frame(canvas, bg=self.CARD_BG)
        win = canvas.create_window((0, 0), window=frm, anchor="nw")
        frm.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        svars = {}
        for col_def in self.columns:
            key  = col_def["key"]
            typ  = col_def.get("type", "entry")
            opts = col_def.get("options", [])
            sv   = tk.StringVar(value=entry.get(key, ""))
            svars[key] = sv

            tk.Label(frm, text=col_def["display"], bg=self.CARD_BG,
                     fg="#555555", font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 0))
            if typ == "combo" and opts:
                w = ttk.Combobox(frm, textvariable=sv, values=opts, width=30)
                self._bind_autocomplete(w, opts, sv)
            elif typ == "combo":
                w = ttk.Combobox(frm, textvariable=sv, width=30)
            else:
                w = ttk.Entry(frm, textvariable=sv, width=32)
            w.pack(fill="x", ipady=2)

        def _save():
            updated = {}
            for col_def in self.columns:
                key = col_def["key"]
                val = svars[key].get().strip()
                if key == "pcba_no":
                    val = val.upper()
                updated[key] = val
            update_db_entry(entry_id, updated)
            self._hist_current_results[idx].update(updated)
            self._populate_hist_tree(self._hist_current_results)
            dlg.destroy()

        btn_bar = tk.Frame(dlg, bg=self.BG, padx=14, pady=8)
        btn_bar.pack(fill="x")
        ttk.Button(btn_bar, text="Save", style="Dark.TButton",
                   command=_save
                   ).pack(side="left", expand=True, fill="x", padx=(0, 6), ipady=5)
        ttk.Button(btn_bar, text="Cancel", style="Muted.TButton",
                   command=dlg.destroy
                   ).pack(side="left", expand=True, fill="x", ipady=5)

    def _hist_delete_entry(self, idx):
        entry   = self._hist_current_results[idx]
        req_key = next((c["key"] for c in self.columns if c.get("required")), None)
        disp    = entry.get(req_key, "") if req_key else ""
        if messagebox.askyesno(
                "Confirm Delete",
                f"Permanently delete entry for '{disp}'?\nThis cannot be undone."):
            delete_db_entry(entry["id"])
            self._hist_current_results.pop(idx)
            self._populate_hist_tree(self._hist_current_results)

    def _hist_delete_selected(self):
        selected = self.hist_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection",
                                "Select one or more rows first.\n"
                                "Use Ctrl+click, Shift+click, or drag to multi-select.")
            return
        n = len(selected)
        if not messagebox.askyesno(
                "Confirm Delete",
                f"Permanently delete {n} selected entr{'y' if n == 1 else 'ies'}?\n"
                "This cannot be undone."):
            return
        # Sort descending so pop() doesn't shift remaining indices
        indices = sorted((int(iid) for iid in selected), reverse=True)
        for idx in indices:
            delete_db_entry(self._hist_current_results[idx]["id"])
            self._hist_current_results.pop(idx)
        self._populate_hist_tree(self._hist_current_results)

    def _hist_export_selected(self):
        selected = self.hist_tree.selection()
        if not selected:
            messagebox.showinfo("No Selection",
                                "Select one or more rows first.\n"
                                "Use Ctrl+click, Shift+click, or drag to multi-select.")
            return
        indices = sorted(int(iid) for iid in selected)
        entries = [self._hist_current_results[i] for i in indices]
        filename = f"SMT_Selected_{date.today().strftime('%Y%m%d')}.xlsx"
        self._do_export(entries, filename)

    # ── session-level actions ─────────────────────────────────────────────────

    def _load_session_as_current(self):
        sel = self.lb_sessions.curselection()
        if not sel:
            messagebox.showinfo("No Session Selected",
                                "Click a session in the list first.")
            return
        session = self._hist_sessions_data[sel[0]]
        if self.session_entries:
            unsaved = len(self.session_entries) - self.session_saved_count
            ans = messagebox.askyesno(
                "Replace Current Session?",
                f"This will replace your current session "
                f"({len(self.session_entries)} entries, {unsaved} unsaved).\n\nContinue?")
            if not ans:
                return
        raw = get_session_entries(session["session_id"])
        # Load raw entries as-is (flexible keys)
        self.session_entries = [{k: v for k, v in e.items() if k not in _META_KEYS}
                                for e in raw]
        self.session_id          = str(uuid.uuid4())
        self.session_saved_count = 0
        # Sync columns to match the loaded session's keys
        self._sync_columns_to_entries(self.session_entries)
        self._refresh_tree()
        self.lbl_counter.config(
            text=f"Entries this session:  {len(self.session_entries)}")
        self.notebook.select(0)
        messagebox.showinfo(
            "Session Loaded",
            f"{len(self.session_entries)} entries loaded as a new editable session.\n"
            "Save Session when done to persist your changes.")

    def _rename_session(self):
        sel = self.lb_sessions.curselection()
        if not sel:
            messagebox.showinfo("No Session Selected", "Click a session first.")
            return
        session = self._hist_sessions_data[sel[0]]
        dlg = tk.Toplevel(self.root); dlg.title("Rename Session")
        dlg.resizable(False, False); dlg.grab_set()
        dlg.configure(bg=self.BG)
        tk.Label(dlg, text="New name:", bg=self.BG,
                 font=("Segoe UI", 9)).grid(row=0, column=0, padx=16, pady=(16, 4), sticky="w")
        sv = tk.StringVar(value=session["label"])
        ent = ttk.Entry(dlg, textvariable=sv, width=32, font=("Segoe UI", 9))
        ent.grid(row=1, column=0, padx=16, pady=(0, 12))
        ent.select_range(0, "end"); ent.focus_set()
        def _save():
            name = sv.get().strip()
            if not name:
                messagebox.showwarning("Empty Name", "Name cannot be blank.", parent=dlg)
                return
            rename_db_session(session["session_id"], name)
            self._hist_refresh_sessions()
            dlg.destroy()
        ent.bind("<Return>", lambda _: _save())
        ttk.Button(dlg, text="Save", style="Dark.TButton",
                   command=_save).grid(row=2, column=0, pady=(0, 14), padx=16, sticky="ew")

    def _delete_session(self):
        sel = self.lb_sessions.curselection()
        if not sel:
            messagebox.showinfo("No Session Selected",
                                "Click a session in the list first.")
            return
        session = self._hist_sessions_data[sel[0]]
        if messagebox.askyesno(
                "Confirm Delete Session",
                f"Permanently delete session:\n\"{session['label']}\"\n\n"
                "All entries in this session will be removed from the database.\n"
                "This cannot be undone."):
            delete_db_session(session["session_id"])
            self._hist_refresh_sessions()
            self._populate_hist_tree([])

    # ── tab switch + window close ─────────────────────────────────────────────

    def _on_tab_change(self, _event):
        tabs = self.notebook.tabs()
        if len(tabs) > 1 and self.notebook.select() == tabs[1]:
            self._hist_refresh_sessions()

    def _on_close(self):
        unsaved = len(self.session_entries) - self.session_saved_count
        if unsaved > 0:
            ans = messagebox.askyesnocancel(
                "Unsaved Entries",
                f"You have {unsaved} unsaved "
                f"entr{'y' if unsaved == 1 else 'ies'} in the current session.\n"
                "Save to database before closing?")
            if ans is None:
                return
            if ans:
                self.save_session()
        self.root.destroy()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
# Quality Sheet DB
# ---------------------------------------------------------------------------

def qs_get_conn():
    return sqlite3.connect(QS_DB_PATH)

def qs_init_db():
    with qs_get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS quality_months (
                id    INTEGER PRIMARY KEY AUTOINCREMENT,
                year  INTEGER NOT NULL,
                month INTEGER NOT NULL,
                UNIQUE(year, month)
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS quality_models (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                month_id       INTEGER NOT NULL,
                sr_no          TEXT    DEFAULT '',
                model          TEXT    NOT NULL,
                line           TEXT    DEFAULT 'LINE 1',
                solder_joints  INTEGER DEFAULT 1,
                dpmo_threshold INTEGER DEFAULT 20,
                FOREIGN KEY (month_id) REFERENCES quality_months(id)
            )""")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS quality_daily (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER NOT NULL,
                day      INTEGER NOT NULL,
                qty      INTEGER DEFAULT 0,
                defects  INTEGER DEFAULT 0,
                line     TEXT    DEFAULT 'LINE 1',
                FOREIGN KEY (model_id) REFERENCES quality_models(id),
                UNIQUE(model_id, day)
            )""")
        # migrate old schema (dpmo_target/opportunities → solder_joints + dpmo_threshold)
        def _cols():
            return [r[1] for r in conn.execute("PRAGMA table_info(quality_models)").fetchall()]
        cols = _cols()
        if "dpmo_target" in cols and "solder_joints" not in cols:
            conn.execute("ALTER TABLE quality_models RENAME COLUMN dpmo_target TO solder_joints")
            cols = _cols()
        if "opportunities" in cols and "solder_joints" not in cols:
            conn.execute("ALTER TABLE quality_models RENAME COLUMN opportunities TO solder_joints")
            cols = _cols()
        if "solder_joints" not in cols:
            conn.execute("ALTER TABLE quality_models ADD COLUMN solder_joints INTEGER DEFAULT 1")
        if "dpmo_threshold" not in cols:
            conn.execute("ALTER TABLE quality_models ADD COLUMN dpmo_threshold INTEGER DEFAULT 20")
        # migrate quality_daily: add missing columns
        dcols = [r[1] for r in conn.execute("PRAGMA table_info(quality_daily)").fetchall()]
        if "line" not in dcols:
            conn.execute("ALTER TABLE quality_daily ADD COLUMN line TEXT DEFAULT 'LINE 1'")
        if "comment" not in dcols:
            conn.execute("ALTER TABLE quality_daily ADD COLUMN comment TEXT DEFAULT ''")
        # migrate quality_months: add optional custom label
        mcols = [r[1] for r in conn.execute("PRAGMA table_info(quality_months)").fetchall()]
        if "label" not in mcols:
            conn.execute("ALTER TABLE quality_months ADD COLUMN label TEXT DEFAULT ''")
        # master model templates
        conn.execute("""
            CREATE TABLE IF NOT EXISTS quality_templates (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                ord            INTEGER DEFAULT 0,
                model          TEXT    NOT NULL,
                solder_joints  INTEGER DEFAULT 1,
                dpmo_threshold INTEGER DEFAULT 20
            )""")
        # add ord column if upgrading from earlier schema
        tcols = [r[1] for r in conn.execute("PRAGMA table_info(quality_templates)").fetchall()]
        if "ord" not in tcols:
            conn.execute("ALTER TABLE quality_templates ADD COLUMN ord INTEGER DEFAULT 0")
            conn.execute("UPDATE quality_templates SET ord=id")
        # seed defaults if empty
        if conn.execute("SELECT COUNT(*) FROM quality_templates").fetchone()[0] == 0:
            # defaults: (model, solder_joints, dpmo_threshold)
            defaults = [
                ("CDA51 TOP", 314, 20), ("CDA51 BOT", 48, 50), ("CDA 63 TOP", 750, 20), ("CDA 63 BOT", 75, 20),
                ("CDD41 TOP", 473, 20), ("CDD41 BOT", 157, 20),
                ("CDD2X", 8, 20), ("CDE87 TOP", 262, 20), ("CDE88 TOP", 262, 20), ("CDE89 TOP", 262, 20),
                ("CDE87 BOT", 53, 20), ("CDE88 BOT", 53, 20), ("CDE89 BOT", 53, 20),
                ("IFP 11", 308, 30), ("CDD81 TOP", 73, 20), ("CDD81 BOT", 40, 20),
                ("CDD91 TOP", 364, 20), ("CDD91 BOT", 424, 20), ("CDD92 TOP", 364, 20), ("CDD92 BOT", 424, 20),
                ("CDV11 TOP", 259, 20), ("CDV11 BOT", 25, 20),
                ("CLS 60", 18, 50), ("CLS 92", 62, 50), ("MCM60 TOP", 828, 20), ("MCM60 BOT", 828, 20),
                ("CDC91 TOP", 556, 50), ("CDC91 BOT", 85, 50), ("CDC81 TOP", 284, 50), ("CDC81 BOT", 251, 50),
                ("CDC82 TOP", 284, 50), ("CDC82 BOT", 251, 50), ("CDC13 TOP", 456, 50),
                ("CDE72", 101, 50), ("CDC22", 8, 50), ("AHM10", 7, 50),
                ("CDE71", 101, 20), ("CDV12 TOP", 259, 20), ("CDV12 BOT", 25, 20),
                ("CDV13 TOP", 259, 20), ("CDV13 BOT", 25, 20),
                ("CLS45", 1, 20), ("MCM71 TOP", 1, 20), ("MCM71 BOT", 1, 20),
                ("CLS 41", 1, 20), ("CLS 44", 1, 20), ("CLS40", 1, 20),
                ("CDA67 TOP", 1, 20), ("CDA67 BOT", 1, 20), ("CDA20 BOT", 1, 20), ("CDA29 TOP", 1, 20),
                ("CDA 11", 750, 20), ("CDA 62 TOP", 75, 20), ("CDC 13 BOT", 59, 20),
                ("CDC 10 TOP", 828, 20), ("CDC10 BOT", 863, 20), ("CDD 94 TOP", 750, 20), ("CDD 94 BOT", 54, 20),
                ("CDA 12", 456, 20), ("CDA 19", 54, 20), ("CDA 29", 364, 20), ("CDA 41 TOP", 424, 20),
                ("CDA 42 TOP", 1, 20),
                ("CDD 93 TOP", 836, 20), ("CDD 93 BOT", 1, 20),
                ("CDA 91", 828, 20), ("CDA92", 828, 20), ("CDA 93", 364, 20), ("CDA68 TOP", 424, 20),
                ("CDV91 TOP", 842, 50), ("CDV91 BOT", 842, 50), ("CDA 14 TOP", 842, 20),
                ("CDA 44 TOP", 828, 20), ("CDA 27 TOP", 1, 20), ("CDA 47 TOP", 828, 20),
                ("CDA 21 TOP", 828, 20), ("CDA 21 BOT", 59, 20), ("CDA 47 BOT", 828, 20),
                ("CDA 44 BOT", 59, 20), ("CDA 90", 842, 20),
            ]
            conn.executemany(
                "INSERT INTO quality_templates(ord,model,solder_joints,dpmo_threshold) VALUES(?,?,?,?)",
                [(i, m, s, t) for i, (m, s, t) in enumerate(defaults, 1)])
        conn.commit()

def qs_get_or_create_month(year, month):
    with qs_get_conn() as conn:
        conn.execute("INSERT OR IGNORE INTO quality_months(year,month) VALUES(?,?)",
                     (year, month))
        conn.commit()
        return conn.execute(
            "SELECT id FROM quality_months WHERE year=? AND month=?",
            (year, month)).fetchone()[0]

def qs_get_all_months():
    """Return all months that have been created, newest first."""
    with qs_get_conn() as conn:
        rows = conn.execute(
            "SELECT id, year, month, label FROM quality_months ORDER BY year DESC, month DESC"
        ).fetchall()
    return [{"id": r[0], "year": r[1], "month": r[2], "label": r[3] or ""} for r in rows]

def qs_rename_month(month_id, label):
    with qs_get_conn() as conn:
        conn.execute("UPDATE quality_months SET label=? WHERE id=?", (label, month_id))
        conn.commit()

def qs_delete_month(month_id):
    with qs_get_conn() as conn:
        model_ids = [r[0] for r in conn.execute(
            "SELECT id FROM quality_models WHERE month_id=?", (month_id,)).fetchall()]
        for mid in model_ids:
            conn.execute("DELETE FROM quality_daily WHERE model_id=?", (mid,))
        conn.execute("DELETE FROM quality_models WHERE month_id=?", (month_id,))
        conn.execute("DELETE FROM quality_months WHERE id=?", (month_id,))
        conn.commit()

def qs_get_templates():
    with qs_get_conn() as conn:
        rows = conn.execute(
            "SELECT id,model,solder_joints,dpmo_threshold FROM quality_templates ORDER BY ord,id"
        ).fetchall()
    return [{"id":r[0],"model":r[1],"solder_joints":r[2],"dpmo_threshold":r[3]} for r in rows]

def qs_add_template(model, solder_joints=1, dpmo_threshold=20):
    with qs_get_conn() as conn:
        max_ord = conn.execute("SELECT COALESCE(MAX(ord),0) FROM quality_templates").fetchone()[0]
        conn.execute("INSERT INTO quality_templates(ord,model,solder_joints,dpmo_threshold) "
                     "VALUES(?,?,?,?)", (max_ord + 1, model, solder_joints, dpmo_threshold))
        conn.commit()

def qs_move_template(tid, direction):
    """Move template up (-1) or down (+1) in ordering."""
    with qs_get_conn() as conn:
        templates = conn.execute(
            "SELECT id, ord FROM quality_templates ORDER BY ord, id").fetchall()
        ids = [r[0] for r in templates]
        if tid not in ids: return
        idx = ids.index(tid)
        swap_idx = idx + direction
        if swap_idx < 0 or swap_idx >= len(ids): return
        id_a, ord_a = templates[idx]
        id_b, ord_b = templates[swap_idx]
        # Swap ord values (use temp to avoid UNIQUE conflicts)
        conn.execute("UPDATE quality_templates SET ord=? WHERE id=?", (ord_b, id_a))
        conn.execute("UPDATE quality_templates SET ord=? WHERE id=?", (ord_a, id_b))
        conn.commit()

def qs_update_template(tid, model, solder_joints, dpmo_threshold):
    with qs_get_conn() as conn:
        conn.execute("UPDATE quality_templates SET model=?,solder_joints=?,dpmo_threshold=? "
                     "WHERE id=?", (model, solder_joints, dpmo_threshold, tid))
        conn.commit()

def qs_delete_template(tid):
    with qs_get_conn() as conn:
        conn.execute("DELETE FROM quality_templates WHERE id=?", (tid,))
        conn.commit()

def qs_populate_month_from_templates(month_id):
    """Insert all template models into a month (used when month has no models yet)."""
    templates = qs_get_templates()
    with qs_get_conn() as conn:
        for i, t in enumerate(templates, 1):
            conn.execute(
                "INSERT INTO quality_models(month_id,sr_no,model,solder_joints,dpmo_threshold) "
                "VALUES(?,?,?,?,?)",
                (month_id, str(i), t["model"], t["solder_joints"], t["dpmo_threshold"]))
        conn.commit()

def qs_get_month_models(month_id):
    with qs_get_conn() as conn:
        rows = conn.execute(
            "SELECT id,sr_no,model,line,solder_joints,dpmo_threshold FROM quality_models "
            "WHERE month_id=? ORDER BY id", (month_id,)).fetchall()
    return [{"id":r[0],"sr_no":r[1],"model":r[2],"line":r[3],
             "solder_joints":r[4],"dpmo_threshold":r[5]}
            for r in rows]

def qs_save_model(month_id, sr_no, model, solder_joints, dpmo_threshold):
    with qs_get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO quality_models(month_id,sr_no,model,solder_joints,dpmo_threshold) "
            "VALUES(?,?,?,?,?)", (month_id, sr_no, model, solder_joints, dpmo_threshold))
        conn.commit()
        return cur.lastrowid

def qs_update_model(model_id, sr_no, model, solder_joints, dpmo_threshold):
    with qs_get_conn() as conn:
        conn.execute(
            "UPDATE quality_models SET sr_no=?,model=?,solder_joints=?,dpmo_threshold=? WHERE id=?",
            (sr_no, model, solder_joints, dpmo_threshold, model_id))
        conn.commit()

def qs_delete_model(model_id):
    with qs_get_conn() as conn:
        conn.execute("DELETE FROM quality_daily  WHERE model_id=?", (model_id,))
        conn.execute("DELETE FROM quality_models WHERE id=?",       (model_id,))
        conn.commit()

def qs_save_daily(model_id, day, qty, defects, line="LINE 1", comment=""):
    with qs_get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO quality_daily(model_id,day,qty,defects,line,comment) "
            "VALUES(?,?,?,?,?,?)", (model_id, day, qty, defects, line, comment or ""))
        conn.commit()

def qs_get_model_daily(model_id):
    with qs_get_conn() as conn:
        rows = conn.execute(
            "SELECT day,qty,defects,line,comment FROM quality_daily WHERE model_id=? ORDER BY day",
            (model_id,)).fetchall()
    return {r[0]: {"qty": r[1], "defects": r[2], "line": r[3] or "LINE 1", "comment": r[4] or ""} for r in rows}

def qs_search(year=None, month=None, model_text=None, day=None):
    with qs_get_conn() as conn:
        q = """SELECT qm.year, qm.month, mo.sr_no, mo.model,
                      mo.solder_joints, mo.dpmo_threshold,
                      d.day, d.qty, d.defects, d.line
               FROM quality_daily d
               JOIN quality_models mo ON mo.id  = d.model_id
               JOIN quality_months qm ON qm.id  = mo.month_id
               WHERE 1=1"""
        params = []
        if year:  q += " AND qm.year=?";  params.append(int(year))
        if month: q += " AND qm.month=?"; params.append(int(month))
        if day:   q += " AND d.day=?";    params.append(int(day))
        q += " ORDER BY qm.year, qm.month, d.day, mo.id"
        rows = conn.execute(q, params).fetchall()
    results = []
    for r in rows:
        yr, mo, sr_no, model, sj, thresh, d, qty, defects, line = r
        if model_text and model_text.upper() not in model.upper():
            continue
        dpmo = (round((defects / (qty * sj)) * 1_000_000, 1)
                if qty and sj else 0)
        results.append({"year":yr,"month":mo,"sr_no":sr_no,"model":model,
                         "line": line or "LINE 1",
                         "solder_joints":sj,"dpmo_threshold":thresh,
                         "day":d,"qty":qty,"defects":defects,"dpmo":dpmo})
    return results

def qs_export_months_multi(month_list, filepath):
    """Export multiple months, each as a separate sheet, into one workbook.
    month_list: list of {"id", "year", "month"}
    """
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    for m in month_list:
        qs_export_month(m["year"], m["month"], m["id"], filepath=None, _wb=wb)
    wb.save(filepath)

def qs_export_month(year, month, month_id, filepath, _wb=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment as XlComment
    from datetime import date as _date

    models  = qs_get_month_models(month_id)
    max_day = calendar.monthrange(year, month)[1]
    if _wb is None:
        wb = Workbook(); wb.remove(wb.active)
        _save = True
    else:
        wb = _wb
        _save = False
    sheet_title = f"{MONTH_NAMES[month-1][:3].upper()}-{str(year)[-2:]}"
    # Avoid duplicate sheet names
    existing = {s.title for s in wb.worksheets}
    suffix = 2
    base = sheet_title
    while sheet_title in existing:
        sheet_title = f"{base}({suffix})"; suffix += 1
    ws = wb.create_sheet(title=sheet_title)

    # ── colour palette ─────────────────────────────────────────────────────
    LINE_COLORS = {"LINE 1": "E67E22", "LINE 2": "2E86C1", "LINE 3": "1E8449",
                   "LINE 4": "C0392B", "LINE 5": "7D3C98"}
    LINE_DEFAULT = "5D6D7E"
    # Light fills for Qty/Defects cells (readable tint matching LINE colour)
    LINE_CELL_FILLS = {
        "LINE 1": PatternFill("solid", fgColor="FDEBD0"),  # pale orange
        "LINE 2": PatternFill("solid", fgColor="D6EAF8"),  # pale blue
        "LINE 3": PatternFill("solid", fgColor="D5F5E3"),  # pale green
        "LINE 4": PatternFill("solid", fgColor="FADBD8"),  # pale red
        "LINE 5": PatternFill("solid", fgColor="E8DAEF"),  # pale purple
    }
    LINE_CELL_DEFAULT = PatternFill("solid", fgColor="E0E0E0")

    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    green_fill = PatternFill("solid", fgColor="27AE60")   # DPMO ≤ threshold (dark green)
    red_fill   = PatternFill("solid", fgColor="C0392B")   # DPMO > threshold (dark red)
    no_fill    = PatternFill("none")                       # plain white

    hdr_font  = Font(bold=True, color="FFFFFF", size=9)
    bold9     = Font(bold=True, size=9)
    bold9_wht = Font(bold=True, size=9, color="FFFFFF")   # white bold for DPMO cells
    norm9     = Font(size=9)
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_mid  = Alignment(horizontal="left",   vertical="center")

    thin      = Side(style="thin",   color="C8C8C8")
    thick     = Side(style="medium", color="1F3864")
    # inner cell border
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    # outer table edge borders (applied to perimeter cells)
    bdr_top   = Border(left=thin, right=thin, top=thick, bottom=thin)
    bdr_bot   = Border(left=thin, right=thin, top=thin,  bottom=thick)
    bdr_left  = Border(left=thick, right=thin, top=thin, bottom=thin)
    bdr_right = Border(left=thin, right=thick, top=thin, bottom=thin)
    bdr_tl    = Border(left=thick, right=thin, top=thick, bottom=thin)
    bdr_tr    = Border(left=thin, right=thick, top=thick, bottom=thin)
    bdr_bl    = Border(left=thick, right=thin, top=thin,  bottom=thick)
    bdr_br    = Border(left=thin, right=thick, top=thin,  bottom=thick)

    # bold TOTAL column font (white text on dark green/red, dark text otherwise)
    total_bold = Font(bold=True, size=9)

    # ── column layout ───────────────────────────────────────────────────────
    # col 1=SR NO., col 2=MODEL, col 3=sub-label, col 4..=days, then 4 summary
    day_start  = 4                      # first day column index
    sum_start  = day_start + max_day    # first summary column index
    total_cols = sum_start + 3          # inclusive last column (4 summary cols)
    SUM_HDRS   = ["TOTAL", "SOLDER\nJOINTS", "HIGHEST\nDPMO DAY", "LOWEST\nDPMO DAY"]

    # ── Row 1: LINE colour tabs + title + LAST UPDATED ──────────────────────
    # Collect distinct lines actually used this month
    all_daily = {m["id"]: qs_get_model_daily(m["id"]) for m in models}
    used_lines_set = set()
    for m in models:
        for v in all_daily[m["id"]].values():
            ln = v.get("line", "LINE 1") or "LINE 1"
            used_lines_set.add(ln)
    # Sort numerically so LINE 1 always appears before LINE 2, etc.
    used_lines_ordered = sorted(used_lines_set,
                                key=lambda s: int(s.split()[-1]) if s.split()[-1].isdigit() else 99)

    tab_col = 1
    for ln in used_lines_ordered:
        col2 = tab_col + 1
        ws.merge_cells(start_row=1, start_column=tab_col,
                       end_row=1,   end_column=col2)
        cell = ws.cell(1, tab_col, ln)
        cell.fill      = PatternFill("solid", fgColor=LINE_COLORS.get(ln, LINE_DEFAULT))
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        tab_col = col2 + 1

    # Title in the middle area (cols 3 → sum_start-1)
    title_start = tab_col
    title_end   = sum_start - 1
    if title_end >= title_start:
        ws.merge_cells(start_row=1, start_column=title_start,
                       end_row=1,   end_column=title_end)
    cell = ws.cell(1, title_start,
                   f"SMT QUALITY REPORT FOR {MONTH_NAMES[month-1].upper()} {year}")
    cell.font      = Font(bold=True, size=11, color="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # LAST UPDATED at far right (merge last 2 cols)
    ws.merge_cells(start_row=1, start_column=sum_start,
                   end_row=1,   end_column=total_cols)
    cell = ws.cell(1, sum_start,
                   f"LAST UPDATED: {_date.today().strftime('%d/%m/%y')}")
    cell.font      = Font(size=8, italic=True, color="555555")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: column headers ───────────────────────────────────────────────
    hdrs = ["SR\nNO.", "MODEL", ""] + \
           [str(d) for d in range(1, max_day + 1)] + SUM_HDRS
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(2, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr
        cell.border = Border(
            left  = thick if c == 1          else thin,
            right = thick if c == total_cols else thin,
            top   = thick,
            bottom= thin)
    ws.row_dimensions[2].height = 36
    # Freeze rows 1-2 (LINE tabs + day headers) and cols 1-3 (SR, MODEL, sub-label)
    ws.freeze_panes = "D3"

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    for d in range(1, max_day + 1):
        ws.column_dimensions[ws.cell(2, day_start + d - 1).column_letter].width = 5.5
    sum_widths = [9, 9, 11, 11]
    for i, w in enumerate(sum_widths):
        ws.column_dimensions[ws.cell(2, sum_start + i).column_letter].width = w

    # ── Data rows ───────────────────────────────────────────────────────────
    row_idx = 3
    for m in models:
        daily  = all_daily[m["id"]]
        sj     = m["solder_joints"]
        thresh = m["dpmo_threshold"]

        days_data = {d: v for d, v in daily.items() if v["qty"] or v["defects"]}
        tqty = sum(v["qty"]     for v in days_data.values())
        tdef = sum(v["defects"] for v in days_data.values())
        overall_dpmo = round((tdef / (tqty * sj)) * 1_000_000, 1) if tqty and sj else 0

        dpmo_per_day = {d: round((v["defects"] / (v["qty"] * sj)) * 1_000_000, 1)
                        for d, v in days_data.items() if v["qty"] and sj}
        max_dpmo_day = max(dpmo_per_day, key=dpmo_per_day.get) if dpmo_per_day else ""
        min_dpmo_day = min(dpmo_per_day, key=dpmo_per_day.get) if dpmo_per_day else ""
        dpmo_label   = f"DPMO ({thresh})"

        is_last_model = (m is models[-1])

        # ── Merged spanning columns: SR NO., MODEL, SOLDER JOINTS, HIGHEST/LOWEST DPMO DAY
        # Merging across 3 sub-rows for visual span.  openpyxl turns non-top-left cells
        # into MergedCell objects (no attributes), so we force real Cell objects back into
        # ws._cells for sub=1 and sub=2 rows to carry the left/right/bottom border data
        # that Excel reads from each individual row in the merge perimeter.
        from openpyxl.cell.cell import Cell as _OpCell

        def _merge_col(col, left_s, right_s, value, font=None, align=None):
            """Merge col across 3 sub-rows and write borders on every row."""
            ws.merge_cells(start_row=row_idx, start_column=col,
                           end_row=row_idx + 2, end_column=col)
            # top-left cell (regular write)
            c = ws.cell(row_idx, col, value)
            c.fill = no_fill
            c.font = font or bold9
            c.alignment = align or ctr
            c.border = Border(left=left_s, right=right_s, top=thin, bottom=thin)
            # sub=1 and sub=2: force real Cells with left/right/bottom borders
            for sub in (1, 2):
                bot_s = thick if (is_last_model and sub == 2) else thin
                fc = _OpCell(ws, row=row_idx + sub, column=col)
                fc.border = Border(left=left_s, right=right_s, top=thin, bottom=bot_s)
                ws._cells[(row_idx + sub, col)] = fc

        _merge_col(1,            thick, thin,  m["sr_no"])
        _merge_col(2,            thin,  thin,  m["model"], align=left_mid)
        _merge_col(sum_start+1,  thin,  thin,  sj)
        _merge_col(sum_start+2,  thin,  thin,  max_dpmo_day or None)
        _merge_col(sum_start+3,  thin,  thick, min_dpmo_day or None)

        # ── Sub-rows: col 3 (sub-label), day cells, TOTAL ────────────────────
        for sub in range(3):
            r   = row_idx + sub
            bot = thick if (is_last_model and sub == 2) else thin

            # Sub-label column (col 3, different label each sub-row)
            sub_lbl = "Qty." if sub == 0 else ("Defects" if sub == 1 else dpmo_label)
            cell = ws.cell(r, 3, sub_lbl)
            cell.fill = no_fill; cell.alignment = left_mid
            cell.font = bold9 if sub == 0 else norm9
            cell.border = Border(left=thin, right=thin, top=thin, bottom=bot)

            # Day value cells
            for d in range(1, max_day + 1):
                dd  = daily.get(d, {})
                q   = dd.get("qty", 0); df = dd.get("defects", 0)
                dln = dd.get("line", "LINE 1") if d in daily else None

                if   sub == 0: val = q  if d in daily else ""
                elif sub == 1: val = df if d in daily else ""
                else:
                    if d not in daily:  val = ""
                    elif q and sj:      val = round((df / (q * sj)) * 1_000_000, 1)
                    else:               val = 0

                cell = ws.cell(r, day_start + d - 1, val)
                cell.alignment = ctr

                if sub == 0 and d in daily:
                    note = dd.get("comment", "")
                    if note:
                        cell.comment = XlComment(note, "SMT")

                if sub == 2:
                    if isinstance(val, (int, float)):
                        cell.fill = red_fill if val > thresh else green_fill
                        cell.font = bold9_wht
                    else:
                        cell.fill = no_fill; cell.font = norm9
                else:
                    cell.fill = (LINE_CELL_FILLS.get(dln, LINE_CELL_DEFAULT)
                                 if dln else no_fill)
                    cell.font = norm9
                cell.border = Border(left=thin, right=thin, top=thin, bottom=bot)

            # TOTAL column (thick left separator, always bold)
            tot_val = tqty if sub == 0 else (tdef if sub == 1 else overall_dpmo)
            cell = ws.cell(r, sum_start, tot_val)
            cell.alignment = ctr
            cell.border = Border(left=thick, right=thin, top=thin, bottom=bot)
            if sub == 2 and isinstance(tot_val, (int, float)):
                cell.fill = red_fill if tot_val > thresh else green_fill
                cell.font = bold9_wht
            else:
                cell.fill = no_fill; cell.font = bold9

        row_idx += 3

    # ── Footer row: PREPARED BY / CHECKED BY ────────────────────────────────
    # Write every cell individually (no merges) so border edges are always
    # on the actual cells Excel reads them from — right border must live on the
    # rightmost cell, bottom border on the bottom row (same row here).
    mid = (total_cols + 1) // 2
    for c in range(1, total_cols + 1):
        val  = "PREPARED BY -" if c == 1 else ("CHECKED BY -" if c == mid else "")
        cell = ws.cell(row_idx, c, val)
        cell.font      = bold9 if val else norm9
        cell.alignment = left_mid
        cell.border    = Border(
            left   = thick if c == 1          else thin,
            right  = thick if c == total_cols else thin,
            top    = thin,
            bottom = thick)
    ws.row_dimensions[row_idx].height = 20

    # ── Single ChartData sheet + per-model Chart sheets ──────────────────────
    def _safe_sheet(base, wb):
        name = base[:31]
        used = {s.title for s in wb.worksheets}
        if name not in used:
            return name
        for i in range(2, 500):
            cand = f"{base[:27]}({i})"[:31]
            if cand not in used:
                return cand
        return name

    # Only models that have at least one day of data
    chart_models = [m for m in models
                    if any(v["qty"] or v["defects"] for v in all_daily[m["id"]].values())]

    if chart_models:
        # ── ChartData sheet ────────────────────────────────────────────────
        # Layout: col A = Day (1-max_day), then per model: Qty col + DPMO col
        cd_title = _safe_sheet(f"{sheet_title} ChartData", wb)
        ws_cd = wb.create_sheet(title=cd_title)

        # Row 1 headers
        ws_cd.cell(1, 1, "Day")
        for mi, m in enumerate(chart_models):
            base_col = 2 + mi * 2
            ws_cd.cell(1, base_col,     f"{m['model']} Qty")
            ws_cd.cell(1, base_col + 1, f"{m['model']} DPMO")

        # Rows 2 … max_day+1: all days, 0 where no data
        for d in range(1, max_day + 1):
            ws_cd.cell(d + 1, 1, d)
            for mi, m in enumerate(chart_models):
                sj_m  = m["solder_joints"]
                v     = all_daily[m["id"]].get(d)
                base_col = 2 + mi * 2
                if v and v["qty"]:
                    dpmo = round((v["defects"] / (v["qty"] * sj_m)) * 1_000_000, 1) \
                           if sj_m else 0
                    ws_cd.cell(d + 1, base_col,     v["qty"])
                    ws_cd.cell(d + 1, base_col + 1, dpmo)
                else:
                    ws_cd.cell(d + 1, base_col,     0)
                    ws_cd.cell(d + 1, base_col + 1, 0)

        # Set column widths so every header is fully visible
        ws_cd.column_dimensions["A"].width = 6   # "Day"
        for mi, m in enumerate(chart_models):
            base_col = 2 + mi * 2
            # header text is "{model} Qty" / "{model} DPMO" — measure the longer one
            w = max(len(f"{m['model']} Qty"), len(f"{m['model']} DPMO")) + 2
            for col in (base_col, base_col + 1):
                ltr = ws_cd.cell(1, col).column_letter
                ws_cd.column_dimensions[ltr].width = w

    if _save:
        wb.save(filepath)

# ---------------------------------------------------------------------------
# Quality Sheet Import
# ---------------------------------------------------------------------------

_IMPORT_MONTH_MAP = {m[:3].upper(): i + 1 for i, m in enumerate(MONTH_NAMES)}

# All known LINE fill colors (both dark tab colors AND light data-cell tints)
_ALL_LINE_FILLS = {
    # dark LINE tab colors (current palette)
    "E67E22": "LINE 1", "2E86C1": "LINE 2", "1E8449": "LINE 3",
    "C0392B": "LINE 4", "7D3C98": "LINE 5",
    # legacy palettes (import compat)
    "F0A500": "LINE 1", "5D6D7E": "LINE 2", "27AE60": "LINE 3",
    "808080": "LINE 2", "3A9D23": "LINE 3", "2471A3": "LINE 4",
    "8E44AD": "LINE 5",
    # light data-cell tints (current)
    "FDEBD0": "LINE 1", "D6EAF8": "LINE 2", "D5F5E3": "LINE 3",
    "FADBD8": "LINE 4", "E8DAEF": "LINE 5",
    # legacy tints
    "FFE8A0": "LINE 1", "D5D8DC": "LINE 2", "DCDCDC": "LINE 2",
    "B7F5CE": "LINE 3", "AED6F1": "LINE 4", "D7BDE2": "LINE 5",
}

def _import_parse_sheet_name(name):
    """Return (year, month) from sheet name like MAR-26, or (None, None)."""
    import re
    m = re.match(r'^([A-Za-z]+)[\s\-_](\d{2,4})', name.strip())
    if m:
        mon = m.group(1).upper()[:3]
        yr  = m.group(2)
        if mon in _IMPORT_MONTH_MAP:
            year = int(yr) if len(yr) == 4 else 2000 + int(yr)
            return year, _IMPORT_MONTH_MAP[mon]
    return None, None

def qs_import_sheet(ws, year, month, overwrite=False):
    """
    Parse one openpyxl worksheet and write data into quality DB.
    Returns a status string.
    """
    import re

    def _v(r, c):
        """Read cell value; returns None for empty/merged cells."""
        try:
            return ws.cell(r, c).value
        except Exception:
            return None

    # ── theme-color palette (read once from the workbook) ─────────────────
    # Standard Office theme slot ordering (matches OOXML spec)
    _THEME_SLOTS = [
        "dk1","lt1","dk2","lt2",
        "accent1","accent2","accent3","accent4","accent5","accent6",
        "hlink","folHlink",
    ]
    _theme_rgb = {}   # slot_index (int) → 6-char hex
    try:
        if ws.parent.loaded_theme:
            from xml.etree import ElementTree as ET
            _NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
            root = ET.fromstring(ws.parent.loaded_theme)
            scheme = root.find(f".//{{{_NS}}}clrScheme")
            if scheme:
                for idx, child in enumerate(scheme):
                    for sub in child:
                        tag = sub.tag.split("}")[-1]
                        if tag == "srgbClr":
                            _theme_rgb[idx] = sub.get("val","000000").upper()
                        elif tag == "sysClr":
                            _theme_rgb[idx] = sub.get("lastClr","000000").upper()
    except Exception:
        pass

    def _tint_rgb(hex6, tint):
        """Apply Excel luminance tint (-1..1) to a 6-char hex color."""
        try:
            r, g, b = int(hex6[0:2],16), int(hex6[2:4],16), int(hex6[4:6],16)
            if tint >= 0:
                r = round(r + (255 - r) * tint)
                g = round(g + (255 - g) * tint)
                b = round(b + (255 - b) * tint)
            else:
                r = round(r * (1 + tint))
                g = round(g * (1 + tint))
                b = round(b * (1 + tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        except Exception:
            return hex6

    def _fill_hex(r, c):
        """Resolve cell fill to a 6-char RGB hex, handling rgb/theme/indexed."""
        try:
            fill = ws.cell(r, c).fill
            if not fill or fill.patternType not in ("solid", "gray125", None):
                # only solid fills carry LINE color info
                pass
            fc = fill.fgColor
            if not fc:
                return None
            ftype = getattr(fc, "type", None)
            if ftype == "rgb":
                rgb = fc.rgb or ""
                # openpyxl prefixes alpha: "FFRRGGBB" → take last 6
                h = rgb[-6:].upper()
                return h if len(h) == 6 and h not in ("000000",) else None
            elif ftype == "theme":
                base = _theme_rgb.get(fc.theme)
                if base:
                    tint = getattr(fc, "tint", 0) or 0
                    return _tint_rgb(base, tint)
            elif ftype == "indexed":
                # Indexed colors (legacy); map most common ones
                _IDX = {
                    0:"000000",1:"FFFFFF",2:"FF0000",3:"00FF00",4:"0000FF",
                    5:"FFFF00",6:"FF00FF",7:"00FFFF",8:"000000",9:"FFFFFF",
                    10:"FF0000",11:"00FF00",12:"0000FF",13:"FFFF00",
                    14:"FF00FF",15:"00FFFF",16:"800000",17:"008000",
                    18:"000080",19:"808000",20:"800080",21:"008080",
                    22:"C0C0C0",23:"808080",24:"9999FF",25:"993366",
                    26:"FFFFCC",27:"CCFFFF",28:"660066",29:"FF8080",
                    30:"0066CC",31:"CCCCFF",32:"000080",33:"FF00FF",
                    34:"FFFF00",35:"00FFFF",36:"800080",37:"800000",
                    38:"008080",39:"0000FF",40:"00CCFF",41:"CCFFFF",
                    42:"CCFFCC",43:"FFFF99",44:"99CCFF",45:"FF99CC",
                    46:"CC99FF",47:"FFCC99",48:"3366FF",49:"33CCCC",
                    50:"99CC00",51:"FFCC00",52:"FF9900",53:"FF6600",
                    54:"666699",55:"969696",56:"003366",57:"339966",
                    58:"003300",59:"333300",60:"993300",61:"993366",
                    62:"333399",63:"333333",64:"000000",65:"FFFFFF",
                }
                return _IDX.get(fc.indexed)
        except Exception:
            pass
        return None

    def _int(v, default=0):
        if v is None or v == "":
            return default
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return default

    def _rgb(hexstr):
        try:
            h = (hexstr or "").strip().lstrip("#")[-6:]
            return int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
        except Exception:
            return None

    def _color_dist(h1, h2):
        a, b = _rgb(h1), _rgb(h2)
        if not a or not b:
            return 999999
        return sum((x-y)**2 for x,y in zip(a,b)) ** 0.5

    # ── Step 1: read LINE tab colors from row 1 ────────────────────────────
    tab_colors  = {}   # "LINE N" → resolved hex
    tab_themes  = {}   # theme_slot (int) → "LINE N"   ← KEY for tint matching
    default_line = "LINE 1"

    for c in range(1, min(ws.max_column or 1, 60) + 1):
        txt = str(_v(1, c) or "").strip().upper()
        if txt.startswith("LINE"):
            try:
                fc = ws.cell(1, c).fill.fgColor
                if getattr(fc, "type", None) == "theme":
                    tab_themes[int(fc.theme)] = txt  # record theme slot
            except Exception:
                pass
            fhex = _fill_hex(1, c)
            if fhex and fhex not in ("000000", "FFFFFF"):
                tab_colors[txt] = fhex

    # Fall back to our known palette if row 1 produced nothing
    if not tab_colors:
        tab_colors = {
            "LINE 1": "F0A500", "LINE 2": "808080", "LINE 3": "3A9D23",
            "LINE 4": "2471A3", "LINE 5": "7D3C98",
        }

    # First LINE present becomes default
    default_line = sorted(tab_colors.keys())[0] if tab_colors else "LINE 1"

    # Known light tints (hex → LINE) for exact-match fallback (both palettes)
    _TINT_TO_LINE = {
        "FDEBD0": "LINE 1", "D5D8DC": "LINE 2", "D5F5E3": "LINE 3",
        "FADBD8": "LINE 4", "E8DAEF": "LINE 5",
        "FFE8A0": "LINE 1", "DCDCDC": "LINE 2", "B7F5CE": "LINE 3",
        "AED6F1": "LINE 4", "D7BDE2": "LINE 5",
    }
    tint_map = {h: ln for h, ln in _TINT_TO_LINE.items() if ln in tab_colors}

    def _nearest_line_by_hex(fhex):
        """Hex-based fallback: exact tint match → nearest tab color."""
        if not fhex or fhex in ("000000", "FFFFFF"):
            return default_line
        if fhex in tint_map:
            return tint_map[fhex]
        if fhex in _ALL_LINE_FILLS and _ALL_LINE_FILLS[fhex] in tab_colors:
            return _ALL_LINE_FILLS[fhex]
        best, best_dist = default_line, 999999
        for line_name, tab_hex in tab_colors.items():
            d = _color_dist(fhex, tab_hex)
            if d < best_dist:
                best_dist, best = d, line_name
        return best

    def _get_cell_line(r, c):
        """Determine LINE for a cell: theme-slot match first (handles tints), then hex distance."""
        try:
            fc = ws.cell(r, c).fill.fgColor
            ftype = getattr(fc, "type", None)
            # Theme index is the most reliable: header + data cells share the same slot
            if ftype == "theme" and int(fc.theme) in tab_themes:
                return tab_themes[int(fc.theme)]
        except Exception:
            pass
        return _nearest_line_by_hex(_fill_hex(r, c))

    # ── Step 2: locate header row ───────────────────────────────────────────
    # Look for a row that has ≥5 integer day numbers (1-31) in it
    header_row  = None
    day_col_map = {}   # day_num -> col_index (1-based)
    sj_col      = None

    for ri in range(1, min(20, ws.max_row or 1) + 1):
        candidate = {}
        for c in range(1, (ws.max_column or 1) + 1):
            raw = _v(ri, c)
            sv  = str(raw or "").strip().replace("\n", " ")
            # Day number cell?
            if isinstance(raw, (int, float)) and 1 <= int(raw) <= 31:
                candidate[int(raw)] = c
            elif sv.isdigit() and 1 <= int(sv) <= 31:
                candidate[int(sv)] = c
            # SOLDER JOINTS header?
            if "SOLDER" in sv.upper() and sj_col is None:
                sj_col = c
        if len(candidate) >= 5:   # found enough day columns
            header_row  = ri
            day_col_map = candidate
            break

    if not header_row or not day_col_map:
        return "skipped — no header row with day columns found (need ≥5 day numbers in one row)"

    # ── Step 3: handle existing month ──────────────────────────────────────
    with qs_get_conn() as conn:
        existing = conn.execute(
            "SELECT id FROM quality_months WHERE year=? AND month=?",
            (year, month)).fetchone()
        existing_id = existing[0] if existing else None

    if existing_id is not None:
        if not overwrite:
            return f"skipped — {MONTH_NAMES[month-1]} {year} already exists (overwrite not selected)"
        # Delete all model rows and daily data for this month
        with qs_get_conn() as conn:
            mids = [r[0] for r in conn.execute(
                "SELECT id FROM quality_models WHERE month_id=?",
                (existing_id,)).fetchall()]
            for mid in mids:
                conn.execute("DELETE FROM quality_daily WHERE model_id=?", (mid,))
            conn.execute("DELETE FROM quality_models WHERE month_id=?", (existing_id,))
            conn.commit()
        month_id = existing_id
    else:
        with qs_get_conn() as conn:
            cur = conn.execute(
                "INSERT INTO quality_months(year, month) VALUES(?,?)", (year, month))
            conn.commit()
            month_id = cur.lastrowid

    # ── Step 4: parse model groups (3 rows per model) ──────────────────────
    imported = 0
    row_idx  = header_row + 1
    max_row  = ws.max_row or 0

    while row_idx + 2 <= max_row:
        # Column 3 of the first sub-row should be "Qty." (or similar)
        sub_lbl = str(_v(row_idx, 3) or "").strip().lower().lstrip("\u2605* ")
        if not (sub_lbl.startswith("qty") or sub_lbl.startswith("quantity")):
            row_idx += 1
            continue

        # SR NO. (col 1) and MODEL (col 2) — both merged across 3 rows
        # so the value lives only in the top (Qty) sub-row
        sr_no = str(_v(row_idx, 1) or "").strip()
        model = str(_v(row_idx, 2) or "").strip()
        if not model:
            row_idx += 1
            continue

        # DPMO threshold from "DPMO (20)" label in 3rd sub-row col 3
        dpmo_lbl = str(_v(row_idx + 2, 3) or "")
        tm       = re.search(r'\((\d+)\)', dpmo_lbl)
        thresh   = int(tm.group(1)) if tm else 20

        # Solder joints — from sj_col, any of the 3 sub-rows
        sj = 1
        if sj_col:
            for sub in range(3):
                v = _v(row_idx + sub, sj_col)
                if v not in (None, ""):
                    try: sj = max(1, int(float(v))); break
                    except (ValueError, TypeError): pass

        # Insert model
        with qs_get_conn() as conn:
            cur = conn.execute(
                "INSERT INTO quality_models(month_id,sr_no,model,solder_joints,dpmo_threshold) "
                "VALUES(?,?,?,?,?)", (month_id, sr_no, model, sj, thresh))
            conn.commit()
            model_id = cur.lastrowid

        # Daily data — determine LINE via theme-slot match (then hex distance)
        for day_num, col in day_col_map.items():
            qty  = _int(_v(row_idx,     col))
            defs = _int(_v(row_idx + 1, col))
            if qty or defs:
                line = _get_cell_line(row_idx, col)
                qs_save_daily(model_id, day_num, qty, defs, line)

        imported += 1
        row_idx += 3

    return f"imported {imported} models → {MONTH_NAMES[month-1]} {year}"


# ---------------------------------------------------------------------------
# Quality Sheet UI
# ---------------------------------------------------------------------------

class QualitySheetFrame(ttk.Frame):
    _LINE_OPTS = ["LINE 1", "LINE 2", "LINE 3", "LINE 4", "LINE 5"]
    _DPMO_OPTS = ["20", "30", "50"]

    def __init__(self, parent, app):
        super().__init__(parent)
        self.app         = app
        self._month_id   = None
        self._year       = datetime.now().year
        self._month      = datetime.now().month
        self._models     = []
        self._sel_model  = None
        self._qs_results = []
        self._build_ui()

    def _build_ui(self):
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True)
        t1 = ttk.Frame(self._nb); self._nb.add(t1, text="  Monthly Entry  ")
        t2 = ttk.Frame(self._nb); self._nb.add(t2, text="  History & Search  ")
        self._build_entry_tab(t1)
        self._build_search_tab(t2)

    # ── Monthly Entry tab ─────────────────────────────────────────────────────

    def _build_entry_tab(self, parent):
        app = self.app
        # header
        hdr = tk.Frame(parent, bg=app.BG, height=50)
        hdr.pack(fill="x", padx=10, pady=(8, 4))
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Year:", bg=app.BG,
                 font=("Segoe UI", 9, "bold")).pack(side="left")
        self.sv_year = tk.StringVar(value=str(self._year))
        tk.Spinbox(hdr, from_=2020, to=2040, textvariable=self.sv_year,
                   width=6, font=("Segoe UI", 9)).pack(side="left", padx=(2, 10))
        tk.Label(hdr, text="Month:", bg=app.BG,
                 font=("Segoe UI", 9, "bold")).pack(side="left")
        self.sv_month = tk.StringVar(value=MONTH_NAMES[self._month - 1])
        ttk.Combobox(hdr, textvariable=self.sv_month, values=MONTH_NAMES,
                     width=11, state="readonly").pack(side="left", padx=(2, 10))
        ttk.Button(hdr, text="Load / New Month", style="Dark.TButton",
                   command=self._load_month).pack(side="left", ipady=4, ipadx=10)
        ttk.Button(hdr, text="Export to Excel", style="Green.TButton",
                   command=self._export_month).pack(side="left", padx=8, ipady=4, ipadx=10)
        ttk.Button(hdr, text="Import from Excel", style="Muted.TButton",
                   command=self._import_from_excel).pack(side="left", ipady=4, ipadx=10)
        ttk.Button(hdr, text="Edit Model Templates", style="Muted.TButton",
                   command=self._open_template_dialog).pack(side="left", padx=8, ipady=4, ipadx=10)
        self.lbl_status = tk.Label(hdr, text="— No month loaded —",
                                   bg=app.BG, fg="#888",
                                   font=("Segoe UI", 9, "italic"))
        self.lbl_status.pack(side="left", padx=10)
        # content
        content = ttk.Frame(parent)
        content.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        # left: model list
        left = tk.Frame(content, bg=app.CARD_BG, width=310)
        left.pack(side="left", fill="y", padx=(0, 6))
        left.pack_propagate(False)
        lhdr = tk.Frame(left, bg=app.DARK_BLUE, height=36)
        lhdr.pack(fill="x"); lhdr.pack_propagate(False)
        tk.Label(lhdr, text="Models / Cards", bg=app.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=10, pady=8)
        # search bar
        sf_outer = tk.Frame(left, bg=app.CARD_BG)
        sf_outer.pack(fill="x", padx=4, pady=(4, 0))
        self.sv_model_search = tk.StringVar()
        self.sv_model_search.trace_add("write", lambda *_: self._refresh_model_list())
        se = ttk.Entry(sf_outer, textvariable=self.sv_model_search)
        se.pack(side="left", fill="x", expand=True, ipady=3)
        ttk.Button(sf_outer, text="✕", width=3,
                   command=lambda: self.sv_model_search.set("")).pack(side="left", padx=(2, 0))
        mf = tk.Frame(left, bg=app.CARD_BG)
        mf.pack(fill="both", expand=True, padx=4, pady=4)
        self.model_tree = ttk.Treeview(mf, columns=("sr","model","sj","thresh"),
                                       show="headings", selectmode="browse")
        for col, hd, w, anch in [("sr","SR",40,"center"),("model","Model",120,"w"),
                                  ("sj","Solder Jts",82,"center"),
                                  ("thresh","DPMO Thr",68,"center")]:
            self.model_tree.heading(col, text=hd)
            self.model_tree.column(col, width=w, minwidth=30, anchor=anch)
        msb = ttk.Scrollbar(mf, orient="vertical", command=self.model_tree.yview)
        self.model_tree.configure(yscrollcommand=msb.set)
        msb.pack(side="right", fill="y")
        self.model_tree.pack(fill="both", expand=True)
        self.model_tree.bind("<<TreeviewSelect>>", self._on_model_select)
        self.model_tree.bind("<Double-1>", self._show_model_graph)
        self.model_tree.tag_configure("odd",  background=app.CARD_BG)
        self.model_tree.tag_configure("even", background=app.ALT_ROW)
        mbtn = tk.Frame(left, bg=app.CARD_BG)
        mbtn.pack(fill="x", padx=4, pady=4)
        ttk.Button(mbtn, text="+ Add",  style="Dark.TButton",
                   command=self._add_model).pack(side="left", ipady=3, ipadx=8)
        ttk.Button(mbtn, text="Edit",   style="Muted.TButton",
                   command=self._edit_model).pack(side="left", padx=4, ipady=3, ipadx=8)
        ttk.Button(mbtn, text="Delete", style="Danger.TButton",
                   command=self._delete_model).pack(side="left", ipady=3, ipadx=8)
        # right: daily entry
        right = tk.Frame(content, bg=app.CARD_BG)
        right.pack(side="left", fill="both", expand=True)
        rhdr = tk.Frame(right, bg=app.DARK_BLUE, height=36)
        rhdr.pack(fill="x"); rhdr.pack_propagate(False)
        self.lbl_daily_title = tk.Label(rhdr, text="Daily Entry  \u2190  select a model",
                                        bg=app.DARK_BLUE, fg="white",
                                        font=("Segoe UI", 10, "bold"))
        self.lbl_daily_title.pack(side="left", padx=10, pady=8)
        tk.Label(rhdr, text="double-click a day to edit",
                 bg=app.DARK_BLUE, fg="#9DB8D8",
                 font=("Segoe UI", 8)).pack(side="right", padx=10)
        df = tk.Frame(right, bg=app.CARD_BG)
        df.pack(fill="both", expand=True, padx=4, pady=4)
        self.day_tree = ttk.Treeview(df, columns=("day","line","qty","defects","dpmo"),
                                     show="headings", selectmode="browse")
        for col, hd, w in [("day","Day",45),("line","Line",72),
                            ("qty","Qty Produced",115),
                            ("defects","Defects",115),("dpmo","DPMO",115)]:
            self.day_tree.heading(col, text=hd)
            self.day_tree.column(col, width=w, anchor="center")
        dsb = ttk.Scrollbar(df, orient="vertical", command=self.day_tree.yview)
        self.day_tree.configure(yscrollcommand=dsb.set)
        dsb.pack(side="right", fill="y")
        self.day_tree.pack(fill="both", expand=True)
        self.day_tree.bind("<Double-1>", self._on_day_double_click)
        self.day_tree.tag_configure("odd",      background=app.CARD_BG)
        self.day_tree.tag_configure("even",     background=app.ALT_ROW)
        self.day_tree.tag_configure("has_data", foreground="#1A5276")
        self.day_tree.tag_configure("defects",  foreground="#922B21")
        self.day_tree.tag_configure("total",    background="#D6EAF8",
                                    foreground="#1A5276", font=("Segoe UI", 9, "bold"))
        # summary strip
        sf = tk.Frame(right, bg="#D6EAF8", height=26)
        sf.pack(fill="x", padx=4, pady=(0, 4)); sf.pack_propagate(False)
        self.lbl_summary = tk.Label(sf, text="", bg="#D6EAF8", fg="#1A5276",
                                    font=("Segoe UI", 9), anchor="w")
        self.lbl_summary.pack(side="left", padx=8)

    # ── Monthly Entry logic ───────────────────────────────────────────────────

    def _load_month(self):
        try:
            year  = int(self.sv_year.get())
            month = MONTH_NAMES.index(self.sv_month.get()) + 1
        except (ValueError, IndexError):
            messagebox.showerror("Invalid", "Please enter a valid year and month.")
            return
        self._year = year; self._month = month
        self._month_id = qs_get_or_create_month(year, month)
        # auto-populate from templates if this month has no models yet
        if not qs_get_month_models(self._month_id):
            qs_populate_month_from_templates(self._month_id)
        self.lbl_status.config(
            text=f"{MONTH_NAMES[month-1]} {year}  (loaded)")
        self._sel_model = None
        self._refresh_model_list()
        self._refresh_day_tree()

    def _refresh_model_list(self):
        self.model_tree.delete(*self.model_tree.get_children())
        if self._month_id is None:
            return
        self._models = qs_get_month_models(self._month_id)
        q = getattr(self, "sv_model_search", None)
        q = q.get().strip().lower() if q else ""
        visible = [m for m in self._models
                   if not q or q in m["model"].lower()]
        for i, m in enumerate(visible):
            tag = "odd" if i % 2 == 0 else "even"
            self.model_tree.insert("", "end", iid=str(m["id"]),
                                   values=(m["sr_no"],m["model"],
                                           m["solder_joints"],m["dpmo_threshold"]),
                                   tags=(tag,))

    def _show_model_graph(self, _event=None):
        sel = self.model_tree.selection()
        if not sel:
            return
        mid = int(sel[0])
        m = next((x for x in self._models if x["id"] == mid), None)
        if m:
            self._open_model_graph(m)

    def _qs_show_model_graph(self, _event=None):
        sel = self.hist_model_tree.selection()
        if not sel:
            return
        mid = int(sel[0])
        m = next((x for x in self._qs_hist_models if x["id"] == mid), None)
        if m:
            self._open_model_graph(m)

    def _open_model_graph(self, model):
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        except ImportError:
            messagebox.showerror("Missing Library",
                                 "matplotlib is required for graphs.\n"
                                 "Install it with:  pip install matplotlib")
            return

        daily  = qs_get_model_daily(model["id"])
        sj     = model["solder_joints"]
        thresh = model["dpmo_threshold"]

        days = sorted(d for d, v in daily.items() if v["qty"] or v["defects"])
        if not days:
            messagebox.showinfo("No Data",
                                f"No daily data recorded for {model['model']}.")
            return

        qty_vals  = [daily[d]["qty"] for d in days]
        dpmo_vals = [
            round((daily[d]["defects"] / (daily[d]["qty"] * sj)) * 1_000_000, 1)
            if daily[d]["qty"] and sj else 0
            for d in days
        ]

        app = self.app
        win = tk.Toplevel(app.root)
        win.title(f"Graph — {model['model']}  (F11 = toggle fullscreen)")
        win.geometry("960x580")
        win.configure(bg=app.BG)
        win.transient(app.root)
        win.state("zoomed")

        # F11 toggles between maximised and normal
        def _toggle_fs(event=None):
            if win.state() == "zoomed":
                win.state("normal")
                win.geometry("960x580")
            else:
                win.state("zoomed")
        win.bind("<F11>", _toggle_fs)

        fig = Figure(figsize=(9.6, 5.4), dpi=96, facecolor=app.BG)
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twinx()

        l1, = ax1.plot(days, qty_vals,  color="#2471A3", marker="o",
                       linewidth=2.2, markersize=6, label="Qty Produced", zorder=3)
        l2, = ax2.plot(days, dpmo_vals, color="#E74C3C", marker="s",
                       linewidth=2.2, markersize=6, linestyle="--", label="DPMO", zorder=3)
        ax2.axhline(thresh, color="#E74C3C", linestyle=":", linewidth=1,
                    alpha=0.55, label=f"Threshold ({thresh})")

        ax1.set_xlabel("Day", fontsize=9)
        ax1.set_ylabel("Qty Produced", color="#2471A3", fontsize=9)
        ax2.set_ylabel("DPMO",         color="#E74C3C",  fontsize=9)
        ax1.tick_params(axis="y", labelcolor="#2471A3")
        ax2.tick_params(axis="y", labelcolor="#E74C3C")
        ax1.set_xticks(days)
        ax1.set_title(f"{model['model']}  —  Daily Production & DPMO",
                      fontsize=11, fontweight="bold", color="#1F3864")
        ax1.grid(axis="y", linestyle=":", alpha=0.4)

        # ── Smart data-point labels ───────────────────────────────────────
        # Each label gets a white background box so graph lines never bleed
        # through the text.  Qty labels alternate above/below based on local
        # slope; DPMO labels go to the opposite side from Qty to minimise
        # overlap between the two series.
        _lbl_box = dict(boxstyle="round,pad=0.18", fc="white", ec="none", alpha=0.85)

        def _slope_sign(vals, i):
            """Return +1 if value is locally rising, -1 if falling."""
            n = len(vals)
            before = vals[i] - vals[i - 1] if i > 0     else 0
            after  = vals[i + 1] - vals[i] if i < n - 1 else 0
            net = before + after
            return 1 if net >= 0 else -1

        for i, (x, y) in enumerate(zip(days, qty_vals)):
            # place above when rising (gives room below for the line coming up),
            # below when falling
            above = _slope_sign(qty_vals, i) >= 0
            dy    = 10 if above else -12
            va    = "bottom" if above else "top"
            ax1.annotate(f"{y:,}", (x, y),
                         textcoords="offset points", xytext=(0, dy),
                         fontsize=9, fontweight="bold", color="#2471A3",
                         ha="center", va=va, bbox=_lbl_box, zorder=5)

        for i, (x, y) in enumerate(zip(days, dpmo_vals)):
            # place on the OPPOSITE side from the qty label at the same day
            qty_above = _slope_sign(qty_vals, i) >= 0 if i < len(qty_vals) else True
            above = not qty_above   # DPMO goes opposite so they don't stack
            dy    = 10 if above else -12
            va    = "bottom" if above else "top"
            ax2.annotate(str(y), (x, y),
                         textcoords="offset points", xytext=(0, dy),
                         fontsize=9, fontweight="bold", color="#C0392B",
                         ha="center", va=va, bbox=_lbl_box, zorder=5)

        lines  = [l1, l2]
        labels = [ln.get_label() for ln in lines]
        fig.legend(lines, labels,
                   loc="lower right",
                   bbox_to_anchor=(0.98, 0.01),
                   fontsize=9, framealpha=0.9, borderpad=0.5)
        fig.tight_layout(pad=1.6, rect=[0, 0.06, 1, 1])

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        tb = NavigationToolbar2Tk(canvas, win)
        tb.update()
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=6, pady=(0, 4))

    def _on_model_select(self, _event=None):
        sel = self.model_tree.selection()
        if not sel:
            return
        mid = int(sel[0])
        self._sel_model = next((m for m in self._models if m["id"] == mid), None)
        if self._sel_model:
            self.lbl_daily_title.config(
                text=f"Daily Entry \u2014 {self._sel_model['model']}  "
                     f"({self._sel_model['line']}, {self._sel_model['solder_joints']} solder joints)")
            self._refresh_day_tree()

    def _refresh_day_tree(self):
        self.day_tree.delete(*self.day_tree.get_children())
        if not self._sel_model or not self._month_id:
            self.lbl_summary.config(text="")
            return
        max_day = calendar.monthrange(self._year, self._month)[1]
        daily   = qs_get_model_daily(self._sel_model["id"])
        sj      = self._sel_model["solder_joints"]
        thresh  = self._sel_model["dpmo_threshold"]
        dpmo_per_day = {}
        for d in range(1, max_day + 1):
            tag  = "odd" if d % 2 != 0 else "even"
            data = daily.get(d)
            if data and (data["qty"] or data["defects"]):
                qty = data["qty"]; defects = data["defects"]
                ln   = data.get("line", "LINE 1")
                dpmo = (round((defects/(qty*sj))*1_000_000, 1) if qty and sj else 0)
                dpmo_per_day[d] = dpmo
                extra = "defects" if dpmo > thresh else "has_data"
                self.day_tree.insert("", "end", iid=str(d),
                                     values=(d, ln, qty, defects, dpmo),
                                     tags=(tag, extra))
            else:
                self.day_tree.insert("", "end", iid=str(d),
                                     values=(d, "\u2014", "\u2014", "\u2014", "\u2014"),
                                     tags=(tag,))
        # TOTAL row + summary strip
        days_data = {d: v for d, v in daily.items() if v["qty"] or v["defects"]}
        if days_data:
            tqty = sum(v["qty"]     for v in days_data.values())
            tdef = sum(v["defects"] for v in days_data.values())
            tot_dpmo = round((tdef/(tqty*sj))*1_000_000, 1) if tqty and sj else 0
            self.day_tree.insert("", "end", iid="total",
                                 values=("TOTAL", "", tqty, tdef, tot_dpmo),
                                 tags=("total",))
            parts = [f"Total  Qty: {tqty:,}  Defects: {tdef:,}  DPMO: {tot_dpmo}",
                     f"Solder Joints: {sj}"]
            if dpmo_per_day:
                hi_d = max(dpmo_per_day, key=dpmo_per_day.get)
                lo_d = min(dpmo_per_day, key=dpmo_per_day.get)
                parts.append(f"Highest DPMO: {dpmo_per_day[hi_d]} (Day {hi_d})")
                parts.append(f"Lowest DPMO: {dpmo_per_day[lo_d]} (Day {lo_d})")
            self.lbl_summary.config(text="    \u2502    ".join(parts))
        else:
            self.lbl_summary.config(text=f"Solder Joints: {sj}  \u2502  No data entered yet")

    def _on_day_double_click(self, event):
        if not self._sel_model:
            return
        item = self.day_tree.identify_row(event.y)
        if item:
            self._open_day_dialog(int(item))

    def _open_day_dialog(self, day):
        app = self.app
        dlg = tk.Toplevel(self); dlg.title(f"Day {day} \u2014 {self._sel_model['model']}")
        dlg.resizable(False, False); dlg.grab_set(); dlg.configure(bg=app.BG)
        data    = qs_get_model_daily(self._sel_model["id"]).get(day, {"qty":0,"defects":0,"line":"LINE 1","comment":""})
        sv_qty  = tk.StringVar(value=str(data["qty"]     or ""))
        sv_def  = tk.StringVar(value=str(data["defects"] or ""))
        sv_line = tk.StringVar(value=data.get("line", "LINE 1"))
        sv_dpmo = tk.StringVar()
        def _calc(*_):
            try:
                q = int(sv_qty.get() or 0); d = int(sv_def.get() or 0)
                sj = self._sel_model["solder_joints"]
                sv_dpmo.set(f"{round((d/(q*sj))*1_000_000,1)}" if q and sj else "0")
            except (ValueError, ZeroDivisionError):
                sv_dpmo.set("\u2014")
        sv_qty.trace_add("write", _calc); sv_def.trace_add("write", _calc); _calc()
        tk.Label(dlg, text=f"Day {day}  \u00b7  {MONTH_NAMES[self._month-1]} {self._year}",
                 bg=app.BG, fg=app.DARK_BLUE,
                 font=("Segoe UI", 10, "bold")).grid(row=0,column=0,columnspan=2,padx=16,pady=(12,6))
        for r,(lbl,widget) in enumerate([
                ("Line:",        ttk.Combobox(dlg, textvariable=sv_line,
                                              values=self._LINE_OPTS, width=13, state="readonly")),
                ("Qty Produced:", tk.Entry(dlg, textvariable=sv_qty,  width=15, font=("Segoe UI",10))),
                ("Defects:",      tk.Entry(dlg, textvariable=sv_def,  width=15, font=("Segoe UI",10))),
                ("DPMO (calc):",  tk.Entry(dlg, textvariable=sv_dpmo, width=15,
                                           font=("Segoe UI",10), state="readonly"))], 1):
            tk.Label(dlg, text=lbl, bg=app.BG,
                     font=("Segoe UI", 9)).grid(row=r,column=0,padx=16,pady=5,sticky="w")
            widget.grid(row=r,column=1,padx=16,pady=5,sticky="w")
        # Comment field
        tk.Label(dlg, text="Comment:", bg=app.BG,
                 font=("Segoe UI", 9)).grid(row=5, column=0, padx=16, pady=5, sticky="nw")
        txt_comment = tk.Text(dlg, width=22, height=3, font=("Segoe UI", 9), wrap="word")
        txt_comment.grid(row=5, column=1, padx=16, pady=5, sticky="w")
        txt_comment.insert("1.0", data.get("comment", ""))
        def _save():
            try:
                qty = int(sv_qty.get() or 0); defects = int(sv_def.get() or 0)
            except ValueError:
                messagebox.showerror("Invalid","Qty and Defects must be whole numbers.",parent=dlg)
                return
            comment = txt_comment.get("1.0", "end-1c").strip()
            qs_save_daily(self._sel_model["id"], day, qty, defects, sv_line.get(), comment)
            self._refresh_day_tree(); dlg.destroy()
        bf = tk.Frame(dlg, bg=app.BG); bf.grid(row=6,column=0,columnspan=2,pady=(6,14))
        ttk.Button(bf, text="Save",   style="Green.TButton",  command=_save).pack(side="left",ipadx=12,ipady=4)
        ttk.Button(bf, text="Cancel", style="Muted.TButton",  command=dlg.destroy).pack(side="left",padx=8,ipadx=12,ipady=4)
        dlg.update_idletasks()
        dlg.geometry(f"+{self.winfo_rootx()+(self.winfo_width()-dlg.winfo_width())//2}"
                     f"+{self.winfo_rooty()+(self.winfo_height()-dlg.winfo_height())//2}")

    def _open_model_dialog(self, model=None):
        app = self.app
        dlg = tk.Toplevel(self)
        dlg.title("Add Model" if model is None else "Edit Model")
        dlg.resizable(False, False); dlg.grab_set(); dlg.configure(bg=app.BG)
        sv_sr    = tk.StringVar(value=model["sr_no"]             if model else "")
        sv_mdl   = tk.StringVar(value=model["model"]              if model else "")
        sv_sj    = tk.StringVar(value=str(model["solder_joints"]) if model else "")
        sv_thresh= tk.StringVar(value=str(model["dpmo_threshold"])if model else "20")
        pad = {"padx":16,"pady":5,"sticky":"w"}
        for r,(lbl,widget) in enumerate([
                ("SR No.:",        tk.Entry(dlg, textvariable=sv_sr,  width=22,font=("Segoe UI",9))),
                ("Model:",         tk.Entry(dlg, textvariable=sv_mdl, width=22,font=("Segoe UI",9))),
                ("Solder Joints:", tk.Entry(dlg, textvariable=sv_sj,  width=22,font=("Segoe UI",9))),
                ("DPMO Threshold\n(for coloring):",
                                   ttk.Combobox(dlg,textvariable=sv_thresh,values=self._DPMO_OPTS,width=20,state="readonly"))]):
            tk.Label(dlg,text=lbl,bg=app.BG,font=("Segoe UI",9)).grid(row=r,column=0,**pad)
            widget.grid(row=r,column=1,**pad)
        def _save():
            mdl = sv_mdl.get().strip()
            if not mdl:
                messagebox.showerror("Required","Model name is required.",parent=dlg); return
            try:    sj = int(sv_sj.get())
            except: messagebox.showerror("Required","Solder Joints must be a number.",parent=dlg); return
            try: thresh = int(sv_thresh.get())
            except ValueError: thresh = 20
            if model is None:
                qs_save_model(self._month_id, sv_sr.get().strip(), mdl, sj, thresh)
            else:
                qs_update_model(model["id"], sv_sr.get().strip(), mdl, sj, thresh)
            self._refresh_model_list(); dlg.destroy()
        bf = tk.Frame(dlg,bg=app.BG); bf.grid(row=4,column=0,columnspan=2,pady=(8,14))
        ttk.Button(bf,text="Save",  style="Green.TButton", command=_save).pack(side="left",ipadx=12,ipady=4)
        ttk.Button(bf,text="Cancel",style="Muted.TButton", command=dlg.destroy).pack(side="left",padx=8,ipadx=12,ipady=4)
        dlg.update_idletasks()
        dlg.geometry(f"+{self.winfo_rootx()+(self.winfo_width()-dlg.winfo_width())//2}"
                     f"+{self.winfo_rooty()+(self.winfo_height()-dlg.winfo_height())//2}")

    # ── Model Template editor ─────────────────────────────────────────────────

    def _open_template_dialog(self):
        app = self.app
        dlg = tk.Toplevel(self)
        dlg.title("Edit Model Templates")
        dlg.grab_set(); dlg.configure(bg=app.BG)
        dlg.resizable(True, True)

        # Header
        tk.Label(dlg, text="Master Model List  —  loaded automatically for every new month",
                 bg=app.BG, fg=app.DARK_BLUE,
                 font=("Segoe UI", 9, "italic")).pack(padx=16, pady=(10, 4), anchor="w")

        # Treeview
        tf = tk.Frame(dlg, bg=app.CARD_BG)
        tf.pack(fill="both", expand=True, padx=12, pady=4)
        tree = ttk.Treeview(tf, columns=("model","sj","thresh"),
                             show="headings", selectmode="browse", height=20)
        for col, hd, w in [("model","Model",200),("sj","Solder Joints",100),
                            ("thresh","DPMO Threshold",120)]:
            tree.heading(col, text=hd); tree.column(col, width=w, anchor="center")
        tree.column("model", anchor="w")
        sb = ttk.Scrollbar(tf, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); tree.pack(fill="both", expand=True)

        def _refresh():
            tree.delete(*tree.get_children())
            for i, t in enumerate(qs_get_templates()):
                tag = "odd" if i % 2 == 0 else "even"
                tree.insert("", "end", iid=str(t["id"]),
                            values=(t["model"], t["solder_joints"], t["dpmo_threshold"]),
                            tags=(tag,))
            tree.tag_configure("odd",  background=app.CARD_BG)
            tree.tag_configure("even", background=app.ALT_ROW)
        _refresh()

        # Buttons
        bf = tk.Frame(dlg, bg=app.BG)
        bf.pack(fill="x", padx=12, pady=(4, 12))

        def _add():
            sub = tk.Toplevel(dlg); sub.title("Add Model")
            sub.grab_set(); sub.configure(bg=app.BG); sub.resizable(False, False)
            sv_m  = tk.StringVar(); sv_sj = tk.StringVar(value="1")
            sv_th = tk.StringVar(value="20")
            pad = {"padx":14,"pady":4,"sticky":"w"}
            for r,(lbl,sv,w) in enumerate([("Model:", sv_m, 22),
                                            ("Solder Joints:", sv_sj, 10),
                                            ("DPMO Threshold:", sv_th, 10)]):
                tk.Label(sub,text=lbl,bg=app.BG,font=("Segoe UI",9)).grid(row=r,column=0,**pad)
                tk.Entry(sub,textvariable=sv,width=w,font=("Segoe UI",9)).grid(row=r,column=1,**pad)
            def _save():
                m = sv_m.get().strip()
                if not m: messagebox.showerror("Required","Model name is required.",parent=sub); return
                try: sj = int(sv_sj.get())
                except: messagebox.showerror("Required","Solder Joints must be a number.",parent=sub); return
                try: th = int(sv_th.get())
                except: th = 20
                qs_add_template(m, sj, th); _refresh(); sub.destroy()
            fb = tk.Frame(sub,bg=app.BG); fb.grid(row=3,column=0,columnspan=2,pady=(6,12))
            ttk.Button(fb,text="Add",   style="Green.TButton", command=_save).pack(side="left",ipadx=10,ipady=4)
            ttk.Button(fb,text="Cancel",style="Muted.TButton", command=sub.destroy).pack(side="left",padx=6,ipadx=10,ipady=4)
            sub.update_idletasks()
            sub.geometry(f"+{dlg.winfo_rootx()+60}+{dlg.winfo_rooty()+60}")

        def _edit():
            sel = tree.selection()
            if not sel: messagebox.showinfo("No Selection","Select a model to edit.",parent=dlg); return
            tid = int(sel[0])
            templates = qs_get_templates()
            t = next((x for x in templates if x["id"] == tid), None)
            if not t: return
            sub = tk.Toplevel(dlg); sub.title("Edit Model")
            sub.grab_set(); sub.configure(bg=app.BG); sub.resizable(False, False)
            sv_m  = tk.StringVar(value=t["model"])
            sv_sj = tk.StringVar(value=str(t["solder_joints"]))
            sv_th = tk.StringVar(value=str(t["dpmo_threshold"]))
            pad = {"padx":14,"pady":4,"sticky":"w"}
            for r,(lbl,sv,w) in enumerate([("Model:", sv_m, 22),
                                            ("Solder Joints:", sv_sj, 10),
                                            ("DPMO Threshold:", sv_th, 10)]):
                tk.Label(sub,text=lbl,bg=app.BG,font=("Segoe UI",9)).grid(row=r,column=0,**pad)
                tk.Entry(sub,textvariable=sv,width=w,font=("Segoe UI",9)).grid(row=r,column=1,**pad)
            def _save():
                m = sv_m.get().strip()
                if not m: messagebox.showerror("Required","Model name is required.",parent=sub); return
                try: sj = int(sv_sj.get())
                except: messagebox.showerror("Required","Solder Joints must be a number.",parent=sub); return
                try: th = int(sv_th.get())
                except: th = 20
                qs_update_template(tid, m, sj, th); _refresh(); sub.destroy()
            fb = tk.Frame(sub,bg=app.BG); fb.grid(row=3,column=0,columnspan=2,pady=(6,12))
            ttk.Button(fb,text="Save",  style="Green.TButton", command=_save).pack(side="left",ipadx=10,ipady=4)
            ttk.Button(fb,text="Cancel",style="Muted.TButton", command=sub.destroy).pack(side="left",padx=6,ipadx=10,ipady=4)
            sub.update_idletasks()
            sub.geometry(f"+{dlg.winfo_rootx()+60}+{dlg.winfo_rooty()+60}")

        def _delete():
            sel = tree.selection()
            if not sel: messagebox.showinfo("No Selection","Select a model to remove.",parent=dlg); return
            vals = tree.item(sel[0])["values"]
            if messagebox.askyesno("Remove Template",
                                   f"Remove \"{vals[0]}\" from the master list?\n"
                                   "(Existing monthly data is not affected.)", parent=dlg):
                qs_delete_template(int(sel[0])); _refresh()

        def _move(direction):
            sel = tree.selection()
            if not sel: return
            tid = int(sel[0])
            qs_move_template(tid, direction)
            _refresh()
            # re-select the moved item
            tree.selection_set(str(tid))
            tree.see(str(tid))

        ttk.Button(bf, text="+ Add",    style="Dark.TButton",
                   command=_add).pack(side="left", ipadx=10, ipady=4)
        ttk.Button(bf, text="Edit",     style="Muted.TButton",
                   command=_edit).pack(side="left", padx=6, ipadx=10, ipady=4)
        ttk.Button(bf, text="Remove",   style="Danger.TButton",
                   command=_delete).pack(side="left", ipadx=10, ipady=4)
        ttk.Button(bf, text="▲ Up",     style="Muted.TButton",
                   command=lambda: _move(-1)).pack(side="left", padx=(12,0), ipadx=8, ipady=4)
        ttk.Button(bf, text="▼ Down",   style="Muted.TButton",
                   command=lambda: _move(1)).pack(side="left", padx=4, ipadx=8, ipady=4)
        ttk.Button(bf, text="Close",    style="Muted.TButton",
                   command=dlg.destroy).pack(side="right", ipadx=10, ipady=4)

        dlg.update_idletasks()
        dlg.geometry(f"460x520+{self.winfo_rootx()+60}+{self.winfo_rooty()+40}")

    def _add_model(self):
        if self._month_id is None:
            messagebox.showinfo("No Month","Load a month first."); return
        self._open_model_dialog()

    def _edit_model(self):
        sel = self.model_tree.selection()
        if not sel: messagebox.showinfo("No Selection","Select a model to edit."); return
        m = next((x for x in self._models if x["id"]==int(sel[0])), None)
        if m: self._open_model_dialog(m)

    def _delete_model(self):
        sel = self.model_tree.selection()
        if not sel: messagebox.showinfo("No Selection","Select a model to delete."); return
        m = next((x for x in self._models if x["id"]==int(sel[0])), None)
        if m and messagebox.askyesno("Confirm Delete",
                f"Delete \"{m['model']}\" and all its daily data?\nCannot be undone."):
            qs_delete_model(m["id"])
            self._sel_model = None
            self._refresh_model_list(); self._refresh_day_tree()

    def _export_month(self):
        if self._month_id is None:
            messagebox.showinfo("No Month","Load a month first."); return
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel workbook","*.xlsx")],
            initialfile=f"Quality_{MONTH_NAMES[self._month-1]}_{self._year}.xlsx",
            title="Save Quality Sheet")
        if not fp: return
        try:
            qs_export_month(self._year, self._month, self._month_id, fp)
            messagebox.showinfo("Exported", f"Saved:\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

    def _import_from_excel(self, filepath=None):
        from openpyxl import load_workbook
        app = self.app
        fp = filepath or filedialog.askopenfilename(
            parent=app.root,
            title="Import Quality Sheet from Excel",
            filetypes=[("Excel workbook", "*.xlsx *.xls")])
        if not fp:
            return
        try:
            wb = load_workbook(fp, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}"); return

        # ── Sheet selection dialog ───────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title("Import from Excel")
        dlg.grab_set(); dlg.configure(bg=app.BG); dlg.resizable(False, False)

        tk.Label(dlg, text="Select sheets to import:", bg=app.BG,
                 font=("Segoe UI", 9, "bold")).pack(padx=16, pady=(12, 4), anchor="w")

        sheet_vars = []
        for sn in wb.sheetnames:
            year, month = _import_parse_sheet_name(sn)
            if year:
                lbl = f"  {sn}   →   {MONTH_NAMES[month-1]} {year}"
                parsed = (year, month)
            else:
                lbl = f"  {sn}   (unrecognized — will be skipped)"
                parsed = None
            var = tk.BooleanVar(value=parsed is not None)
            cb  = tk.Checkbutton(dlg, text=lbl, variable=var, bg=app.BG,
                                  font=("Segoe UI", 9),
                                  state="normal" if parsed else "disabled")
            cb.pack(padx=16, anchor="w")
            sheet_vars.append((sn, var, parsed))

        ttk.Separator(dlg, orient="horizontal").pack(fill="x", padx=16, pady=8)

        overwrite_var = tk.BooleanVar(value=False)
        tk.Checkbutton(dlg, text="Overwrite existing data if month already exists",
                       variable=overwrite_var, bg=app.BG,
                       font=("Segoe UI", 9)).pack(padx=16, anchor="w")

        result = {"go": False}
        def _go(): result["go"] = True; dlg.destroy()

        bf = tk.Frame(dlg, bg=app.BG)
        bf.pack(pady=(10, 14))
        ttk.Button(bf, text="Import", style="Green.TButton",
                   command=_go).pack(side="left", ipadx=12, ipady=4)
        ttk.Button(bf, text="Cancel", style="Muted.TButton",
                   command=dlg.destroy).pack(side="left", padx=8, ipadx=12, ipady=4)
        dlg.update_idletasks()
        dlg.geometry(
            f"+{self.winfo_rootx()+(self.winfo_width()-dlg.winfo_width())//2}"
            f"+{self.winfo_rooty()+(self.winfo_height()-dlg.winfo_height())//2}")
        dlg.wait_window()

        if not result["go"]:
            return

        selected = [(sn, p) for sn, var, p in sheet_vars if var.get() and p]
        if not selected:
            messagebox.showinfo("Nothing Selected",
                                "No valid sheets were selected."); return

        overwrite = overwrite_var.get()
        msgs = []
        for sn, (year, month) in selected:
            ws = wb[sn]
            try:
                msg = qs_import_sheet(ws, year, month, overwrite)
            except Exception as e:
                msg = f"error — {e}"
            msgs.append(f"{sn}: {msg}")

        messagebox.showinfo("Import Complete", "\n".join(msgs))
        self._qs_refresh_sessions()
        # Reload current view if an imported month matches the loaded month
        if self._month_id is not None:
            for _, (yr, mo) in selected:
                if yr == self._year and mo == self._month:
                    self._load_month()
                    break

    # ── History & Search tab ──────────────────────────────────────────────────

    def _build_search_tab(self, parent):
        app = self.app
        body = tk.Frame(parent, bg=app.BG)
        body.pack(fill="both", expand=True, padx=10, pady=(8, 6))

        # ── Left: sessions list ──────────────────────────────────────────────
        left = tk.Frame(body, bg=app.CARD_BG, width=190)
        left.pack(side="left", fill="y", padx=(0, 6))
        left.pack_propagate(False)

        lhdr = tk.Frame(left, bg=app.DARK_BLUE, height=34)
        lhdr.pack(fill="x"); lhdr.pack_propagate(False)
        tk.Label(lhdr, text="Sessions (Months)", bg=app.DARK_BLUE, fg="white",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=10, pady=7)

        lf = tk.Frame(left, bg=app.CARD_BG)
        lf.pack(fill="both", expand=True, padx=4, pady=4)
        self.sess_lb = tk.Listbox(lf, selectmode="extended",
                                  font=("Segoe UI", 9), activestyle="none",
                                  bg=app.CARD_BG, fg="#1F2F40",
                                  selectbackground=app.DARK_BLUE,
                                  selectforeground="white",
                                  relief="flat", borderwidth=0)
        lb_sb = ttk.Scrollbar(lf, orient="vertical", command=self.sess_lb.yview)
        self.sess_lb.configure(yscrollcommand=lb_sb.set)
        lb_sb.pack(side="right", fill="y")
        self.sess_lb.pack(fill="both", expand=True)
        self.sess_lb.bind("<<ListboxSelect>>", self._qs_on_session_select)

        lbtn = tk.Frame(left, bg=app.CARD_BG)
        lbtn.pack(fill="x", padx=4, pady=(0, 4))
        ttk.Button(lbtn, text="Open as Current",  style="Dark.TButton",
                   command=self._qs_open_as_current).pack(fill="x", ipady=4)
        ttk.Button(lbtn, text="Export Selected",  style="Green.TButton",
                   command=self._qs_export_selected).pack(fill="x", pady=(4,0), ipady=4)
        ttk.Button(lbtn, text="Rename Session",   style="Muted.TButton",
                   command=self._qs_rename_session).pack(fill="x", pady=(4,0), ipady=3)
        ttk.Button(lbtn, text="Delete Session",   style="Danger.TButton",
                   command=self._qs_delete_session).pack(fill="x", pady=(4,0), ipady=3)
        ttk.Button(lbtn, text="Refresh Sessions", style="Muted.TButton",
                   command=self._qs_refresh_sessions).pack(fill="x", pady=(4,0), ipady=3)

        # ── Right: switcher frame ─────────────────────────────────────────────
        self._qs_right = tk.Frame(body, bg=app.BG)
        self._qs_right.pack(side="left", fill="both", expand=True)

        # ── Right panel A: session month view (single selection) ──────────────
        self._qs_month_frame = tk.Frame(self._qs_right, bg=app.BG)
        # models sub-panel
        mv_left = tk.Frame(self._qs_month_frame, bg=app.CARD_BG, width=310)
        mv_left.pack(side="left", fill="y", padx=(0, 6))
        mv_left.pack_propagate(False)
        mvlhdr = tk.Frame(mv_left, bg=app.DARK_BLUE, height=36)
        mvlhdr.pack(fill="x"); mvlhdr.pack_propagate(False)
        tk.Label(mvlhdr, text="Models / Cards", bg=app.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=10, pady=8)
        mvlf = tk.Frame(mv_left, bg=app.CARD_BG)
        mvlf.pack(fill="both", expand=True, padx=4, pady=4)
        self.hist_model_tree = ttk.Treeview(mvlf,
                                            columns=("sr","model","sj","thresh"),
                                            show="headings", selectmode="browse")
        for col, hd, w, anch in [("sr","SR",40,"center"),("model","Model",120,"w"),
                                  ("sj","Solder Jts",82,"center"),
                                  ("thresh","DPMO Thr",68,"center")]:
            self.hist_model_tree.heading(col, text=hd)
            self.hist_model_tree.column(col, width=w, minwidth=30, anchor=anch)
        hmsb = ttk.Scrollbar(mvlf, orient="vertical", command=self.hist_model_tree.yview)
        self.hist_model_tree.configure(yscrollcommand=hmsb.set)
        hmsb.pack(side="right", fill="y")
        self.hist_model_tree.pack(fill="both", expand=True)
        self.hist_model_tree.bind("<<TreeviewSelect>>", self._qs_on_hist_model_select)
        self.hist_model_tree.bind("<Double-1>", self._qs_show_model_graph)
        self.hist_model_tree.tag_configure("odd",  background=app.CARD_BG)
        self.hist_model_tree.tag_configure("even", background=app.ALT_ROW)
        # days sub-panel
        mv_right = tk.Frame(self._qs_month_frame, bg=app.CARD_BG)
        mv_right.pack(side="left", fill="both", expand=True)
        mvrhdr = tk.Frame(mv_right, bg=app.DARK_BLUE, height=36)
        mvrhdr.pack(fill="x"); mvrhdr.pack_propagate(False)
        self.lbl_hist_title = tk.Label(mvrhdr,
                                       text="Daily Data  \u2190  select a model",
                                       bg=app.DARK_BLUE, fg="white",
                                       font=("Segoe UI", 10, "bold"))
        self.lbl_hist_title.pack(side="left", padx=10, pady=8)
        mvdf = tk.Frame(mv_right, bg=app.CARD_BG)
        mvdf.pack(fill="both", expand=True, padx=4, pady=4)
        self.hist_day_tree = ttk.Treeview(mvdf,
                                          columns=("day","line","qty","defects","dpmo"),
                                          show="headings", selectmode="browse")
        for col, hd, w in [("day","Day",45),("line","Line",72),
                            ("qty","Qty Produced",115),
                            ("defects","Defects",115),("dpmo","DPMO",115)]:
            self.hist_day_tree.heading(col, text=hd)
            self.hist_day_tree.column(col, width=w, anchor="center")
        hdsb = ttk.Scrollbar(mvdf, orient="vertical", command=self.hist_day_tree.yview)
        self.hist_day_tree.configure(yscrollcommand=hdsb.set)
        hdsb.pack(side="right", fill="y")
        self.hist_day_tree.pack(fill="both", expand=True)
        self.hist_day_tree.tag_configure("odd",      background=app.CARD_BG)
        self.hist_day_tree.tag_configure("even",     background=app.ALT_ROW)
        self.hist_day_tree.tag_configure("has_data", foreground="#1A5276")
        self.hist_day_tree.tag_configure("defects",  foreground="#922B21")
        # summary strip
        hsf = tk.Frame(mv_right, bg="#D6EAF8", height=26)
        hsf.pack(fill="x", padx=4, pady=(0, 4)); hsf.pack_propagate(False)
        self.lbl_hist_summary = tk.Label(hsf, text="", bg="#D6EAF8", fg="#1A5276",
                                         font=("Segoe UI", 9), anchor="w")
        self.lbl_hist_summary.pack(side="left", padx=8)

        # ── Right panel B: flat multi-session table ───────────────────────────
        self._qs_flat_frame = tk.Frame(self._qs_right, bg=app.BG)
        # Filter row
        frow = tk.Frame(self._qs_flat_frame, bg=app.BG)
        frow.pack(fill="x", pady=(0, 4))
        def _lbl(t):
            tk.Label(frow, text=t, bg=app.BG,
                     font=("Segoe UI", 9, "bold")).pack(side="left")
        _lbl("Model:")
        self.sv_s_model = tk.StringVar()
        tk.Entry(frow, textvariable=self.sv_s_model, width=18,
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 10))
        _lbl("Day:")
        self.sv_s_day = tk.StringVar()
        tk.Entry(frow, textvariable=self.sv_s_day, width=4,
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 10))
        ttk.Button(frow, text="Filter", style="Dark.TButton",
                   command=self._qs_filter).pack(side="left", ipady=4, ipadx=10)
        ttk.Button(frow, text="Clear",  style="Muted.TButton",
                   command=self._qs_clear).pack(side="left", padx=6, ipady=4, ipadx=8)
        rf = tk.Frame(self._qs_flat_frame, bg=app.CARD_BG)
        rf.pack(fill="both", expand=True)
        rh = tk.Frame(rf, bg=app.DARK_BLUE, height=34)
        rh.pack(fill="x"); rh.pack_propagate(False)
        tk.Label(rh, text="Results", bg=app.DARK_BLUE, fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=10, pady=7)
        self.lbl_qs_count = tk.Label(rh, text="", bg=app.DARK_BLUE, fg="#9DB8D8",
                                     font=("Segoe UI", 9))
        self.lbl_qs_count.pack(side="right", padx=10)
        cols = ("month_yr","sr_no","model","line","sj","dpmo_thresh",
                "day","date","qty","defects","dpmo")
        self.qs_tree = ttk.Treeview(rf, columns=cols, show="headings",
                                    selectmode="extended")
        for col, hd, w in [("month_yr","Month/Year",100),("sr_no","SR",45),
                            ("model","Model",120),("line","Line",65),
                            ("sj","Solder Jts",80),("dpmo_thresh","DPMO Thr",75),
                            ("day","Day",42),("date","Date",90),
                            ("qty","Qty",72),("defects","Defects",72),("dpmo","DPMO",72)]:
            self.qs_tree.heading(col, text=hd)
            self.qs_tree.column(col, width=w, anchor="center")
        self.qs_tree.column("model", anchor="w")
        vsb = ttk.Scrollbar(rf, orient="vertical",   command=self.qs_tree.yview)
        hsb = ttk.Scrollbar(rf, orient="horizontal", command=self.qs_tree.xview)
        self.qs_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side="bottom", fill="x"); vsb.pack(side="right", fill="y")
        self.qs_tree.pack(fill="both", expand=True)
        self.qs_tree.tag_configure("odd",  background=app.CARD_BG)
        self.qs_tree.tag_configure("even", background=app.ALT_ROW)

        # initialise state
        self._qs_results        = []
        self._qs_all_months     = []
        self._qs_loaded_months  = []
        self._qs_hist_models    = []   # models for single-session view
        self._qs_hist_month     = None # the single month dict shown in month view
        self._qs_refresh_sessions()
        # default: show flat frame
        self._qs_flat_frame.pack(fill="both", expand=True)

    # ── sessions helpers ──────────────────────────────────────────────────────

    def _qs_rename_session(self):
        idxs = self.sess_lb.curselection()
        if len(idxs) != 1:
            messagebox.showinfo("Select One Session",
                                "Select exactly one session to rename.")
            return
        m = self._qs_all_months[idxs[0]]
        default = f"{MONTH_NAMES[m['month']-1]}  {m['year']}"
        current  = m["label"] if m["label"] else default
        dlg = tk.Toplevel(self); dlg.title("Rename Session")
        dlg.resizable(False, False); dlg.grab_set()
        dlg.configure(bg=app.BG)
        tk.Label(dlg, text="New name:", bg=app.BG,
                 font=("Segoe UI", 9)).grid(row=0, column=0, padx=16, pady=(16,4), sticky="w")
        sv = tk.StringVar(value=current)
        ent = ttk.Entry(dlg, textvariable=sv, width=32, font=("Segoe UI", 9))
        ent.grid(row=1, column=0, padx=16, pady=(0,12))
        ent.select_range(0, "end"); ent.focus_set()
        def _save():
            name = sv.get().strip()
            if not name:
                messagebox.showwarning("Empty Name", "Name cannot be blank.", parent=dlg)
                return
            # Saving the default label as empty so it stays auto-named
            qs_rename_month(m["id"], "" if name == default else name)
            self._qs_refresh_sessions()
            dlg.destroy()
        ent.bind("<Return>", lambda _: _save())
        ttk.Button(dlg, text="Save", style="Dark.TButton",
                   command=_save).grid(row=2, column=0, pady=(0,14), padx=16, sticky="ew")

    def _qs_delete_session(self):
        idxs = self.sess_lb.curselection()
        if not idxs:
            messagebox.showinfo("No Session Selected",
                                "Select a session to delete.")
            return
        months = [self._qs_all_months[i] for i in idxs]
        names  = "\n".join(
            m["label"] if m["label"] else f"{MONTH_NAMES[m['month']-1]} {m['year']}"
            for m in months)
        if not messagebox.askyesno(
                "Confirm Delete",
                f"Permanently delete {len(months)} session(s):\n{names}\n\n"
                "All model and daily data will be removed.\nThis cannot be undone."):
            return
        for m in months:
            qs_delete_month(m["id"])
        self._qs_refresh_sessions()
        # clear displayed data
        for w in self._qs_right.winfo_children():
            w.pack_forget()

    def _qs_refresh_sessions(self):
        self._qs_all_months = qs_get_all_months()
        self.sess_lb.delete(0, "end")
        for m in self._qs_all_months:
            default = f"{MONTH_NAMES[m['month']-1]}  {m['year']}"
            self.sess_lb.insert("end", m["label"] if m["label"] else default)

    def _qs_on_session_select(self, _event=None):
        idxs = self.sess_lb.curselection()
        if not idxs:
            return
        self._qs_loaded_months = [self._qs_all_months[i] for i in idxs]
        if len(idxs) == 1:
            # Single session → show month view
            self._qs_flat_frame.pack_forget()
            self._qs_month_frame.pack(fill="both", expand=True)
            self._qs_hist_month = self._qs_loaded_months[0]
            self._qs_load_hist_month(self._qs_hist_month)
        else:
            # Multiple → show flat table
            self._qs_month_frame.pack_forget()
            self._qs_flat_frame.pack(fill="both", expand=True)
            results = []
            for m in self._qs_loaded_months:
                results += qs_search(year=m["year"], month=m["month"])
            self._qs_results = results
            self._qs_apply_filter()

    def _qs_load_hist_month(self, mdict):
        """Populate the month-view model tree for a single selected session."""
        self._qs_hist_models = qs_get_month_models(mdict["id"])
        self.hist_model_tree.delete(*self.hist_model_tree.get_children())
        for i, m in enumerate(self._qs_hist_models):
            tag = "odd" if i % 2 == 0 else "even"
            self.hist_model_tree.insert("", "end", iid=str(m["id"]),
                                        values=(m["sr_no"], m["model"],
                                                m["solder_joints"], m["dpmo_threshold"]),
                                        tags=(tag,))
        self.lbl_hist_title.config(text="Daily Data  \u2190  select a model")
        self.hist_day_tree.delete(*self.hist_day_tree.get_children())
        self.lbl_hist_summary.config(text="")

    def _qs_on_hist_model_select(self, _event=None):
        sel = self.hist_model_tree.selection()
        if not sel or not self._qs_hist_month:
            return
        mid = int(sel[0])
        m = next((x for x in self._qs_hist_models if x["id"] == mid), None)
        if not m:
            return
        mdict = self._qs_hist_month
        self.lbl_hist_title.config(
            text=f"Daily Data \u2014 {m['model']}  ({m['solder_joints']} solder joints)")
        self.hist_day_tree.delete(*self.hist_day_tree.get_children())
        max_day = calendar.monthrange(mdict["year"], mdict["month"])[1]
        daily   = qs_get_model_daily(mid)
        sj      = m["solder_joints"]
        thresh  = m["dpmo_threshold"]
        dpmo_per_day = {}
        for d in range(1, max_day + 1):
            tag  = "odd" if d % 2 != 0 else "even"
            data = daily.get(d)
            if data and (data["qty"] or data["defects"]):
                qty = data["qty"]; defects = data["defects"]
                ln   = data.get("line", "LINE 1")
                dpmo = round((defects/(qty*sj))*1_000_000, 1) if qty and sj else 0
                dpmo_per_day[d] = dpmo
                extra = "defects" if dpmo > thresh else "has_data"
                self.hist_day_tree.insert("", "end", iid=str(d),
                                          values=(d, ln, qty, defects, dpmo),
                                          tags=(tag, extra))
            else:
                self.hist_day_tree.insert("", "end", iid=str(d),
                                          values=(d, "\u2014", "\u2014", "\u2014", "\u2014"),
                                          tags=(tag,))
        days_data = {d: v for d, v in daily.items() if v["qty"] or v["defects"]}
        if days_data:
            tqty = sum(v["qty"]     for v in days_data.values())
            tdef = sum(v["defects"] for v in days_data.values())
            tot_dpmo = round((tdef/(tqty*sj))*1_000_000, 1) if tqty and sj else 0
            parts = [f"Total  Qty: {tqty:,}  Defects: {tdef:,}  DPMO: {tot_dpmo}",
                     f"Solder Joints: {sj}"]
            if dpmo_per_day:
                hi_d = max(dpmo_per_day, key=dpmo_per_day.get)
                lo_d = min(dpmo_per_day, key=dpmo_per_day.get)
                parts.append(f"Highest DPMO: {dpmo_per_day[hi_d]} (Day {hi_d})")
                parts.append(f"Lowest DPMO: {dpmo_per_day[lo_d]} (Day {lo_d})")
            self.lbl_hist_summary.config(text="    \u2502    ".join(parts))
        else:
            self.lbl_hist_summary.config(text="No data for this model.")

    def _qs_open_as_current(self):
        """Load the single selected session into the Monthly Entry tab."""
        idxs = self.sess_lb.curselection()
        if not idxs:
            messagebox.showinfo("No Selection",
                                "Select a single month to open as current session."); return
        if len(idxs) > 1:
            messagebox.showinfo("Multiple Selected",
                                "Select only one month to open as current session."); return
        m = self._qs_all_months[idxs[0]]
        self.sv_year.set(str(m["year"]))
        self.sv_month.set(MONTH_NAMES[m["month"] - 1])
        self._load_month()
        self._nb.select(0)  # switch to Monthly Entry tab

    def _qs_filter(self):
        self._qs_apply_filter()

    def _qs_apply_filter(self):
        mt  = self.sv_s_model.get().strip().upper()
        ds  = self.sv_s_day.get().strip()
        day = int(ds) if ds.isdigit() else None
        filtered = [r for r in self._qs_results
                    if (not mt  or mt  in r["model"].upper()) and
                       (not day or day == r["day"])]
        self._populate_qs_tree(filtered)

    def _qs_clear(self):
        self.sv_s_model.set(""); self.sv_s_day.set("")
        self._populate_qs_tree(self._qs_results)

    def _populate_qs_tree(self, results):
        self.qs_tree.delete(*self.qs_tree.get_children())
        for i, r in enumerate(results):
            try:    ds = datetime(r["year"], r["month"], r["day"]).strftime("%Y-%m-%d")
            except: ds = f"{r['year']}-{r['month']:02d}-{r['day']:02d}"
            tag = "odd" if i % 2 == 0 else "even"
            self.qs_tree.insert("", "end", iid=str(i), tags=(tag,),
                values=(f"{MONTH_NAMES[r['month']-1][:3]} {r['year']}",
                        r["sr_no"], r["model"], r["line"],
                        r["solder_joints"], r["dpmo_threshold"],
                        r["day"], ds, r["qty"], r["defects"], r["dpmo"]))
        n = len(results)
        self.lbl_qs_count.config(text=f"{n} row{'s' if n!=1 else ''}")

    def _qs_export_selected(self):
        idxs = self.sess_lb.curselection()
        if not idxs:
            messagebox.showinfo("No Selection",
                                "Select one or more months in the Sessions list."); return
        months = [self._qs_all_months[i] for i in idxs]
        label = "_".join(f"{MONTH_NAMES[m['month']-1][:3]}{m['year']}" for m in months[:3])
        if len(months) > 3: label += "_etc"
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook","*.xlsx")],
            initialfile=f"QS_{label}.xlsx",
            title="Export Selected Months")
        if not fp: return
        try:
            qs_export_months_multi(months, fp)
            names = ", ".join(f"{MONTH_NAMES[m['month']-1]} {m['year']}" for m in months)
            messagebox.showinfo("Exported",
                                f"Saved {len(months)} sheet(s):\n{names}\n\n{fp}")
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc))

# ---------------------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app  = SMTReworkApp(root)
    root.mainloop()
